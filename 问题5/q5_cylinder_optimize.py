# -*- coding: utf-8 -*-
"""
问题5：圆柱严格全遮蔽下，5 架无人机每架至多 3 枚烟幕干扰弹，干扰 M1/M2/M3
目标：三枚导弹有效遮蔽时长之和 最大化

基于配对矩阵的快速分派 + warm-start:
  M1 ← FY1 (多弹, Q3 warm-start)
  M2 ← FY2 + FY4 + FY3
  M3 ← FY5
"""
from __future__ import annotations

import json
import shutil
import sys
import time
from pathlib import Path

import numpy as np

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
from common.smoke_geom import (  # noqa: E402
    DROP_GAP,
    MISSILES,
    PTS_FAST,
    PTS_FINE,
    PTS_OPT,
    UAVS,
    _MISSILE_CACHE,
    device_info,
    heading_deg,
    kinematics,
    kinematics_batch,
    shield_duration,
    union_duration_batch,
    union_duration_single_path,
)

OUT = ROOT / "结果"

def display_path(path):
    return Path(path).resolve().relative_to(ROOT.parent.resolve()).as_posix()

OUT.mkdir(exist_ok=True)
ATTACH = ROOT.parent / "附件"

MODE = "cylinder"
PTS_COARSE = PTS_FAST
PTS_MID = PTS_OPT
PTS_HI = PTS_FINE

UAV_NAMES = ["FY1", "FY2", "FY3", "FY4", "FY5"]
MIS_NAMES = ["M1", "M2", "M3"]

# probe / Q2-Q4 已知好种子: (theta, v, t_drop, tau)
WARM_SEEDS = {
    ("FY1", "M1"): [
        [3.0825785063, 70.0, 0.0, 2.4841194504],  # Q2/Q4
        [3.13544, 139.99828, 0.00298, 3.61113],  # Q3 bomb1
        [3.129, 140.0, 0.0, 3.511],
        [0.1511, 70.0, 0.0, 3.0],  # mid-run th≈8.66°
    ],
    ("FY2", "M1"): [[5.34404, 136.85, 8.551, 3.986]],
    ("FY2", "M2"): [
        [5.035, 136.1, 5.60, 1.95],
        [3.963, 134.6, 5.0, 3.0],
        [3.96292, 134.6266, 4.319, 5.804],  # prev run
    ],
    ("FY3", "M1"): [[2.13758, 92.96, 31.733, 7.646]],
    ("FY3", "M2"): [[1.515, 133.4, 24.29, 0.30], [1.515, 133.4, 24.0, 1.5]],
    ("FY4", "M2"): [[4.147, 81.6, 12.54, 11.86], [4.147, 90.0, 12.0, 10.0]],
    ("FY5", "M1"): [[1.780, 136.7, 12.55, 2.64]],
    ("FY5", "M3"): [
        [2.010, 140.0, 12.44, 0.30],
        [2.010, 140.0, 12.0, 1.0],
        [2.010, 130.0, 10.0, 2.0],
        [2.010, 140.0, 11.0, 1.5],
        [2.05, 140.0, 12.5, 0.5],
    ],
}

# Q3 multi-bomb schedule for FY1-M1
Q3_MULTI = {
    "theta": 3.13544,
    "v": 139.99828,
    "bombs": [
        (0.00298, 3.61113),
        (3.70246, 5.33749),
        (5.56948, 6.04053),
    ],
}


def aim_theta(fy0, target_xy=(0.0, 0.0)):
    dxy = np.array(target_xy, dtype=float) - np.asarray(fy0, dtype=float)[:2]
    return float(np.arctan2(dxy[1], dxy[0]))


def mis_info(name):
    m0 = MISSILES[name]
    u, v, t_hit, dist = _MISSILE_CACHE[name]
    return m0, v, t_hit, dist


def highres_missile(bombs, missile, pts=None, n=150000):
    if pts is None:
        pts = PTS_HI
    if not bombs:
        return 0.0, [], []
    total, per, segs = union_duration_single_path(
        bombs, mode=MODE, pts=pts, n=n, refine=True, missile=missile
    )
    return total, per, segs


def bomb_info(uav_name, missile, th, v, td, tau, bomb_id=1):
    fy0 = UAVS[uav_name]
    v_fy, p_drop, t_det, p_det = kinematics(th, v, td, tau, fy0=fy0)
    dur, segs = shield_duration(
        p_det, t_det, mode=MODE, pts=PTS_MID, n=60000, refine=True, missile=missile
    )
    return {
        "uav": uav_name,
        "missile": missile,
        "theta": float(th),
        "heading_deg": heading_deg(th),
        "v": float(v),
        "t_drop": float(td),
        "tau": float(tau),
        "t_det": float(t_det),
        "P_drop": p_drop.tolist(),
        "P_det": p_det.tolist(),
        "duration": float(dur),
        "segments": segs,
        "bomb_id": bomb_id,
    }


def optimize_single_bomb_for_uav_missile(
    uav_name,
    missile,
    n_rand=1800,
    n_top=5,
    use_gpu=True,
    seed_extra=0,
):
    fy0 = UAVS[uav_name]
    m0, v_m, t_hit, _ = mis_info(missile)
    th0 = aim_theta(fy0)
    th_m = aim_theta(fy0, target_xy=m0[:2])
    th_t = aim_theta(fy0, target_xy=(0.0, 200.0))
    rng = np.random.default_rng(
        (abs(hash((uav_name, missile))) + seed_extra * 9973) % (2**32)
    )

    seeds = list(WARM_SEEDS.get((uav_name, missile), []))
    seeds.extend(
        [
            [th0, 120, 0.0, 3.0],
            [th0, 140, 0.0, 2.5],
            [th0, 100, 1.0, 4.0],
            [th_m, 120, 1.0, 4.0],
            [th_m, 140, 0.0, 3.0],
            [th_t, 110, 2.0, 5.0],
            [th_t, 140, 1.0, 3.5],
        ]
    )

    ths = (
        [th0 + d for d in np.linspace(-1.0, 1.0, 9)]
        + [th_m + d for d in np.linspace(-0.6, 0.6, 5)]
        + [th_t + d for d in np.linspace(-0.6, 0.6, 5)]
    )
    # include warm thetas
    for s in seeds:
        ths.append(s[0])
    vs = [70, 90, 110, 130, 140]
    tds = list(np.linspace(0.0, min(45.0, t_hit * 0.85), 12))
    taus = [0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 6.0, 8.0, 10.0, 12.0]
    grid = []
    for th in ths:
        for v in vs:
            for td in tds[::2]:
                for tau in taus[::2]:
                    grid.append([th, v, td, tau])

    Xr = np.column_stack(
        [
            (th0 + (rng.random(n_rand) - 0.5) * 2.2) % (2 * np.pi),
            70 + rng.random(n_rand) * 70,
            rng.random(n_rand) * min(45.0, t_hit * 0.85),
            0.4 + rng.random(n_rand) * 12,
        ]
    )
    Xi = []
    for _ in range(600):
        th = th0 + (rng.random() - 0.5) * 1.0
        v = 70 + rng.random() * 70
        t_det_g = 2.0 + rng.random() * min(40.0, t_hit * 0.75)
        tau = 0.5 + rng.random() * 10
        td = max(0.0, t_det_g - tau)
        Xi.append([th, v, td, tau])

    X = np.vstack(
        [
            np.asarray(seeds, dtype=float),
            np.asarray(grid, dtype=float),
            Xr,
            np.asarray(Xi, dtype=float),
        ]
    )
    X[:, 0] = np.mod(X[:, 0], 2 * np.pi)
    X[:, 1] = np.clip(X[:, 1], 70, 140)
    X[:, 2] = np.clip(X[:, 2], 0, 50)
    X[:, 3] = np.clip(X[:, 3], 0.3, 16)
    X = np.unique(np.round(X, 5), axis=0)

    th, v, td, tau = X[:, 0], X[:, 1], X[:, 2], X[:, 3]
    _, pdet, tdet = kinematics_batch(th, v, td, tau, fy0=fy0)
    durs, _ = union_duration_batch(
        pdet[:, None, :],
        tdet[:, None],
        mode=MODE,
        pts=PTS_COARSE,
        n_time=3000,
        use_gpu=use_gpu,
        missile=missile,
    )
    order = np.argsort(-durs)
    best_x = X[order[0]].copy()
    best_d = float(durs[order[0]])

    if best_d <= 0:
        info = bomb_info(uav_name, missile, *best_x, bomb_id=1)
        info["duration"] = 0.0
        info["segments"] = []
        return 0.0, best_x, info

    lb = np.array([0.0, 70.0, 0.0, 0.3])
    ub = np.array([2 * np.pi, 140.0, 50.0, 16.0])
    for k in range(min(n_top, len(order))):
        if durs[order[k]] <= 0:
            break
        x = X[order[k]].copy()
        span = np.array([0.08, 8.0, 1.2, 1.2])
        val = float(durs[order[k]])
        for it in range(18):
            improved = False
            for d in range(4):
                for sgn in (-1.0, 1.0):
                    trial = x.copy()
                    trial[d] += sgn * span[d]
                    trial[0] = trial[0] % (2 * np.pi)
                    trial = np.clip(trial, lb, ub)
                    _, pd, td_ = kinematics_batch(
                        [trial[0]], [trial[1]], [trial[2]], [trial[3]], fy0=fy0
                    )
                    dv, _ = union_duration_batch(
                        pd[:, None, :],
                        td_[:, None],
                        mode=MODE,
                        pts=PTS_MID,
                        n_time=3500,
                        use_gpu=use_gpu,
                        missile=missile,
                    )
                    if dv[0] > val + 1e-9:
                        val = float(dv[0])
                        x = trial
                        improved = True
            noise = (rng.random((20, 4)) - 0.5) * 2 * span
            trials = np.clip(x[None, :] + noise, lb, ub)
            trials[:, 0] %= 2 * np.pi
            _, pd, td_ = kinematics_batch(
                trials[:, 0], trials[:, 1], trials[:, 2], trials[:, 3], fy0=fy0
            )
            dvs, _ = union_duration_batch(
                pd[:, None, :],
                td_[:, None],
                mode=MODE,
                pts=PTS_MID,
                n_time=3000,
                use_gpu=use_gpu,
                missile=missile,
            )
            j = int(np.argmax(dvs))
            if dvs[j] > val + 1e-9:
                val = float(dvs[j])
                x = trials[j]
                improved = True
            span *= 0.9 if improved else 1.04
            span = np.minimum(span, np.array([0.12, 12, 2.0, 2.0]))
        if val > best_d:
            best_d = val
            best_x = x

    info = bomb_info(uav_name, missile, *best_x, bomb_id=1)
    return float(info["duration"]), best_x, info


def optimize_multi_bombs_uav(
    uav_name,
    missile,
    n_bombs=2,
    base_x=None,
    use_gpu=True,
    also_search_heading=False,
):
    """同一航向/速度下优化多弹；默认不扫大量 (θ,v) 以加速。"""
    fy0 = UAVS[uav_name]
    rng = np.random.default_rng(abs(hash((uav_name, missile, n_bombs))) % (2**32))
    if base_x is None:
        _, base_x, _ = optimize_single_bomb_for_uav_missile(
            uav_name, missile, n_rand=1000, n_top=3, use_gpu=use_gpu
        )

    candidates_tv = [(float(base_x[0]), float(base_x[1]))]
    if also_search_heading:
        th0 = float(base_x[0])
        v0 = float(base_x[1])
        for dth in [-0.08, 0.0, 0.08]:
            for dv in [-15, 0, 15]:
                candidates_tv.append((th0 + dth, float(np.clip(v0 + dv, 70, 140))))
    if uav_name == "FY1" and missile == "M1":
        candidates_tv.append((Q3_MULTI["theta"], Q3_MULTI["v"]))
        candidates_tv.append((3.0825785063, 70.0))
        candidates_tv.append((np.pi, 140.0))

    # unique
    seen = set()
    uniq = []
    for th, v in candidates_tv:
        key = (round(th, 5), round(v, 3))
        if key not in seen:
            seen.add(key)
            uniq.append((th, v))
    candidates_tv = uniq

    best_total = -1.0
    best_bombs = []

    for th, v in candidates_tv:
        dim = 2 * n_bombs
        lb = np.array([0.0, 0.3] * n_bombs)
        ub = np.array([40.0, 14.0] * n_bombs)

        def fix(x):
            x = np.asarray(x, dtype=float).copy()
            ts = []
            for i in range(n_bombs):
                t = max(0.0, x[2 * i])
                if i > 0:
                    t = max(t, ts[-1] + DROP_GAP)
                ts.append(t)
                x[2 * i] = t
                x[2 * i + 1] = max(0.3, min(16.0, x[2 * i + 1]))
            return x

        def eval_x(xs, n_time=3000, pts=PTS_COARSE):
            xs = np.atleast_2d(xs)
            N = xs.shape[0]
            dets, tds = [], []
            for i in range(n_bombs):
                td = xs[:, 2 * i]
                tau = xs[:, 2 * i + 1]
                ths = np.full(N, th)
                vs = np.full(N, v)
                _, pd, td_ = kinematics_batch(ths, vs, td, tau, fy0=fy0)
                dets.append(pd)
                tds.append(td_)
            P = np.stack(dets, axis=1)
            T = np.stack(tds, axis=1)
            dur, per = union_duration_batch(
                P, T, mode=MODE, pts=pts, n_time=n_time, use_gpu=use_gpu, missile=missile
            )
            return dur, per

        rows = []
        base_td, base_tau = float(base_x[2]), float(base_x[3])
        if n_bombs == 2:
            for dt in [1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0, 8.0]:
                for dtau in [-1.0, 0.0, 1.0, 2.0, 3.0]:
                    rows.append(
                        [base_td, base_tau, base_td + dt, max(0.5, base_tau + dtau)]
                    )
            if uav_name == "FY1" and missile == "M1":
                rows.append([0.003, 3.611, 3.702, 5.337])
                rows.append([0.0, 2.5, 3.5, 5.0])
                rows.append([0.0, 3.5, 4.0, 5.5])
        else:
            for dt1 in [1.2, 2.0, 3.0, 4.0]:
                for dt2 in [1.2, 2.0, 3.5, 5.0]:
                    rows.append(
                        [
                            base_td,
                            base_tau,
                            base_td + dt1,
                            max(0.5, base_tau + 1.0),
                            base_td + dt1 + dt2,
                            max(0.5, base_tau + 2.0),
                        ]
                    )
            if uav_name == "FY1" and missile == "M1":
                rows.append([0.003, 3.611, 3.702, 5.337, 5.569, 6.041])
                rows.append([0.0, 2.5, 3.5, 4.5, 6.0, 5.5])
                rows.append([0.0, 3.0, 2.5, 4.0, 5.0, 5.0])

        n_rand = 900 if n_bombs == 2 else 700
        Xr = lb + rng.random((n_rand, dim)) * (ub - lb)
        X = np.vstack([np.asarray(rows, dtype=float), Xr])
        X = np.array([fix(r) for r in X])

        durs, _ = eval_x(X, n_time=3000, pts=PTS_COARSE)
        order = np.argsort(-durs)
        x_best = X[order[0]].copy()
        best = float(durs[order[0]])

        span = np.array([1.0, 1.0] * n_bombs)
        for it in range(16):
            improved = False
            for d in range(dim):
                for sgn in (-1.0, 1.0):
                    trial = fix(np.clip(x_best + sgn * span * np.eye(1, dim, d).ravel(), lb, ub))
                    val = float(eval_x(trial[None, :], n_time=3500, pts=PTS_MID)[0][0])
                    if val > best + 1e-9:
                        best = val
                        x_best = trial
                        improved = True
            noise = (rng.random((18, dim)) - 0.5) * 2 * span
            trials = np.array([fix(np.clip(x_best + n, lb, ub)) for n in noise])
            vals = eval_x(trials, n_time=3000, pts=PTS_MID)[0]
            j = int(np.argmax(vals))
            if vals[j] > best + 1e-9:
                best = float(vals[j])
                x_best = trials[j]
                improved = True
            span *= 0.9 if improved else 1.04

        bombs = []
        for i in range(n_bombs):
            td = float(x_best[2 * i])
            tau = float(x_best[2 * i + 1])
            bombs.append(bomb_info(uav_name, missile, th, v, td, tau, bomb_id=i + 1))
        bomb_pairs = [(np.array(b["P_det"]), b["t_det"]) for b in bombs]
        total, per, segs = highres_missile(bomb_pairs, missile, pts=PTS_MID, n=60000)
        for i, b in enumerate(bombs):
            b["duration"] = float(per[i]) if i < len(per) else b["duration"]
        if total > best_total:
            best_total = total
            best_bombs = bombs

    return best_total, best_bombs


def seed_q3_multi():
    """直接用 Q3 圆柱三弹方案作为 FY1→M1 候选。"""
    th, v = Q3_MULTI["theta"], Q3_MULTI["v"]
    bombs = []
    for i, (td, tau) in enumerate(Q3_MULTI["bombs"], 1):
        bombs.append(bomb_info("FY1", "M1", th, v, td, tau, bomb_id=i))
    pairs = [(np.array(b["P_det"]), b["t_det"]) for b in bombs]
    total, per, segs = highres_missile(pairs, "M1", pts=PTS_MID, n=80000)
    for i, b in enumerate(bombs):
        b["duration"] = float(per[i]) if i < len(per) else b["duration"]
    return total, bombs


def write_result3_xlsx(all_bombs, path: Path, summary=None):
    template = ATTACH / "result3.xlsx"
    if load_workbook is not None and template.exists():
        wb = load_workbook(template)
        ws = wb.active
    else:
        if Workbook is None:
            raise RuntimeError("openpyxl required")
        wb = Workbook()
        ws = wb.active
        headers = [
            "无人机编号",
            "无人机运动方向",
            "无人机运动速度 (m/s)",
            "烟幕干扰弹编号",
            "烟幕干扰弹投放点的x坐标 (m)",
            "烟幕干扰弹投放点的y坐标 (m)",
            "烟幕干扰弹投放点的z坐标 (m)",
            "烟幕干扰弹起爆点的x坐标 (m)",
            "烟幕干扰弹起爆点的y坐标 (m)",
            "烟幕干扰弹起爆点的z坐标 (m)",
            "有效干扰时长 (s)",
            "干扰的导弹编号",
        ]
        for j, h in enumerate(headers, 1):
            ws.cell(1, j, h)

    by_uav = {n: [] for n in UAV_NAMES}
    for b in all_bombs:
        by_uav[b["uav"]].append(b)

    r = 2
    for name in UAV_NAMES:
        bombs = sorted(by_uav[name], key=lambda x: x.get("bomb_id", 1))
        if bombs:
            heading = bombs[0]["heading_deg"]
            v = bombs[0]["v"]
        else:
            heading = None
            v = None
        for bid in range(1, 4):
            ws.cell(r, 1, name)
            ws.cell(r, 4, bid)
            matched = [b for b in bombs if b.get("bomb_id", 1) == bid]
            if not matched and bid <= len(bombs):
                matched = [bombs[bid - 1]]
            if matched:
                b = matched[0]
                ws.cell(r, 2, float(b["heading_deg"]))
                ws.cell(r, 3, float(b["v"]))
                for j in range(3):
                    ws.cell(r, 5 + j, float(b["P_drop"][j]))
                    ws.cell(r, 8 + j, float(b["P_det"][j]))
                ws.cell(r, 11, float(b.get("duration", 0.0)))
                ws.cell(r, 12, b.get("missile", ""))
            else:
                if heading is not None:
                    ws.cell(r, 2, float(heading))
                    ws.cell(r, 3, float(v))
            r += 1

    if summary:
        ws.cell(r + 1, 1, summary)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    try:
        shutil.copy2(path, ATTACH / "result3.xlsx")
    except Exception as e:
        print("copy to 附件 failed:", e)


def choose_best_plan(uav, missile, use_gpu=True, try_triple=True, try_double=True):
    d1, x1, info1 = optimize_single_bomb_for_uav_missile(
        uav, missile, n_rand=1800, n_top=5, use_gpu=use_gpu
    )
    print(
        f"  {uav} 单弹 vs {missile}: {d1:.4f}s  "
        f"th={info1['heading_deg']:.2f} v={info1['v']:.1f} "
        f"td={info1['t_drop']:.2f} tau={info1['tau']:.2f}"
    )
    candidates = [(d1, [info1] if d1 > 1e-9 else [])]

    # Q3 multi seed for FY1-M1
    if uav == "FY1" and missile == "M1":
        dq, bq = seed_q3_multi()
        print(f"  FY1 Q3三弹种子 vs M1: {dq:.4f}s")
        if dq > 0:
            candidates.append((dq, bq))

    if d1 > 1e-6 and try_double:
        d2, bombs2 = optimize_multi_bombs_uav(
            uav,
            missile,
            n_bombs=2,
            base_x=x1,
            use_gpu=use_gpu,
            also_search_heading=(uav == "FY1"),
        )
        print(f"  {uav} 双弹 vs {missile}: {d2:.4f}s")
        if d2 > 0:
            candidates.append((d2, bombs2))

    if d1 > 1e-6 and try_triple:
        d3, bombs3 = optimize_multi_bombs_uav(
            uav,
            missile,
            n_bombs=3,
            base_x=x1,
            use_gpu=use_gpu,
            also_search_heading=(uav == "FY1"),
        )
        print(f"  {uav} 三弹 vs {missile}: {d3:.4f}s")
        if d3 > 0:
            candidates.append((d3, bombs3))

    candidates = [(d, b) for d, b in candidates if b]
    if not candidates:
        return 0.0, []
    candidates.sort(key=lambda t: -t[0])
    best_d, best_bombs = candidates[0]
    best_bombs = [b for b in best_bombs if b.get("duration", 0) > 1e-6] or best_bombs[:1]
    # renumber
    for i, b in enumerate(best_bombs, 1):
        b["bomb_id"] = i
    print(f"  => 选用 {len(best_bombs)} 弹, 机内并集≈{best_d:.4f}s")
    return best_d, best_bombs


def main():
    t0 = time.perf_counter()
    use_gpu = True
    print("==== 问题5 圆柱严格全遮蔽 5机 vs M1/M2/M3 (快速 warm-start) ====")
    print("device:", device_info())
    for m in MIS_NAMES:
        _, _, th, dist = mis_info(m)
        print(f"  {m}: t_hit={th:.4f} dist={dist:.2f}")

    assignment = {
        "M1": ["FY1"],
        "M2": ["FY2", "FY4", "FY3"],
        "M3": ["FY5"],
    }
    print("分派方案:", assignment)

    missile_details = {}
    missile_totals = {}
    used_uavs = set()

    for missile, uavs in assignment.items():
        print(f"\n---- 优化 {missile} 使用 {uavs} ----")
        bombs_for_m = []
        for ui, uav in enumerate(uavs):
            if uav in used_uavs:
                continue
            # 主无人机可三弹；副机单弹优先（FY2 可双弹）
            try_triple = ui == 0 and missile in ("M1", "M3")
            try_double = ui == 0 or (missile == "M2" and uav == "FY2")
            if missile == "M2" and uav in ("FY3", "FY4"):
                try_double = False
                try_triple = False
            _, bombs = choose_best_plan(
                uav,
                missile,
                use_gpu=use_gpu,
                try_triple=try_triple,
                try_double=try_double,
            )
            if bombs and any(b.get("duration", 0) > 1e-9 for b in bombs):
                bombs_for_m.extend(bombs)
                used_uavs.add(uav)
            else:
                print(f"  {uav} 对 {missile} 无有效遮蔽")

        pairs = [(np.array(b["P_det"]), b["t_det"]) for b in bombs_for_m]
        total_m, per_m, segs_m = highres_missile(pairs, missile, pts=PTS_HI, n=180000)
        for i, b in enumerate(bombs_for_m):
            if i < len(per_m):
                b["duration"] = float(per_m[i])
        bombs_for_m = [b for b in bombs_for_m if b.get("duration", 0) > 1e-9] or bombs_for_m
        missile_totals[missile] = total_m
        missile_details[missile] = {
            "segments": segs_m,
            "per": list(per_m),
            "bombs": bombs_for_m,
        }
        print(f"  {missile} 总并集 = {total_m:.6f} s  segs={segs_m}")

    # 未使用机补强
    unused = [u for u in UAV_NAMES if u not in used_uavs]
    viable = {
        "FY1": ["M1"],
        "FY2": ["M1", "M2"],
        "FY3": ["M1", "M2"],
        "FY4": ["M2"],
        "FY5": ["M1", "M3"],
    }
    if unused:
        print(f"\n未使用无人机 {unused}，按可配对补强")
        for uav in unused:
            cands = viable.get(uav, MIS_NAMES)
            target = min(cands, key=lambda m: missile_totals.get(m, 0.0))
            _, bombs = choose_best_plan(
                uav, target, use_gpu=use_gpu, try_triple=False, try_double=False
            )
            if not bombs or all(b.get("duration", 0) <= 1e-9 for b in bombs):
                for alt in cands:
                    if alt == target:
                        continue
                    _, bombs = choose_best_plan(
                        uav, alt, use_gpu=use_gpu, try_triple=False, try_double=False
                    )
                    if bombs and any(b.get("duration", 0) > 1e-9 for b in bombs):
                        target = alt
                        break
            bombs = [b for b in bombs if b.get("duration", 0) > 1e-9]
            if not bombs:
                print(f"  {uav} 无有效遮蔽，跳过")
                continue
            old = missile_details[target]["bombs"]
            new_bombs = old + bombs
            pairs = [(np.array(b["P_det"]), b["t_det"]) for b in new_bombs]
            total_m, per_m, segs_m = highres_missile(pairs, target, pts=PTS_HI, n=180000)
            for i, b in enumerate(new_bombs):
                if i < len(per_m):
                    b["duration"] = float(per_m[i])
            new_bombs = [b for b in new_bombs if b.get("duration", 0) > 1e-9]
            missile_totals[target] = total_m
            missile_details[target] = {
                "segments": segs_m,
                "per": list(per_m),
                "bombs": new_bombs,
            }
            used_uavs.add(uav)
            print(f"  {uav} 补强 {target} -> 并集 {total_m:.6f}s")

    # 组装 all bombs，统一同机航向速度
    all_bombs = []
    for m in MIS_NAMES:
        all_bombs.extend(missile_details[m]["bombs"])

    by_uav = {n: [] for n in UAV_NAMES}
    for b in all_bombs:
        by_uav[b["uav"]].append(b)
    all_bombs_final = []
    for name in UAV_NAMES:
        blist = sorted(by_uav[name], key=lambda x: x["t_drop"])
        if not blist:
            continue
        th0, v0 = blist[0]["theta"], blist[0]["v"]
        for i, b in enumerate(blist[:3]):
            b = dict(b)
            b["bomb_id"] = i + 1
            if abs(b["theta"] - th0) > 1e-9 or abs(b["v"] - v0) > 1e-9:
                v_fy, p_drop, t_det, p_det = kinematics(
                    th0, v0, b["t_drop"], b["tau"], fy0=UAVS[name]
                )
                b["theta"] = th0
                b["heading_deg"] = heading_deg(th0)
                b["v"] = v0
                b["P_drop"] = p_drop.tolist()
                b["P_det"] = p_det.tolist()
                b["t_det"] = float(t_det)
            all_bombs_final.append(b)

    all_bombs = all_bombs_final

    # 终算各导弹并集
    missile_totals = {}
    missile_details = {}
    for m in MIS_NAMES:
        bombs_m = [b for b in all_bombs if b["missile"] == m]
        pairs = [(np.array(b["P_det"]), b["t_det"]) for b in bombs_m]
        total_m, per_m, segs_m = highres_missile(pairs, m, pts=PTS_HI, n=200000)
        for i, b in enumerate(bombs_m):
            if i < len(per_m):
                b["duration"] = float(per_m[i])
        missile_totals[m] = total_m
        missile_details[m] = {
            "segments": segs_m,
            "per": list(per_m),
            "n_bombs": len(bombs_m),
        }
        print(f"终算 {m}: union={total_m:.6f} s, n_bombs={len(bombs_m)}, segs={segs_m}")

    total_sum = float(sum(missile_totals.values()))
    elapsed = time.perf_counter() - t0
    print(f"\n==== 问题5 汇总 ====")
    print(f"M1+M2+M3 总有效遮蔽 = {total_sum:.6f} s")
    for m in MIS_NAMES:
        print(f"  {m}: {missile_totals[m]:.6f} s")
    print(f"总弹数 = {len(all_bombs)}")

    def conv(o):
        if isinstance(o, (np.floating,)):
            return float(o)
        if isinstance(o, (np.integer,)):
            return int(o)
        if isinstance(o, np.ndarray):
            return o.tolist()
        if isinstance(o, list):
            return [conv(x) for x in o]
        if isinstance(o, tuple):
            return [conv(x) for x in o]
        if isinstance(o, dict):
            return {k: conv(v) for k, v in o.items()}
        return o

    result = {
        "total_sum": total_sum,
        "missile_totals": missile_totals,
        "missile_details": missile_details,
        "bombs": all_bombs,
        "mode": "cylinder_strict_full_cover",
        "device": device_info(),
        "elapsed_s": elapsed,
        "assignment": assignment,
    }
    with open(OUT / "q5_cylinder_result.json", "w", encoding="utf-8") as f:
        json.dump(conv(result), f, ensure_ascii=False, indent=2)

    with open(OUT / "q5_cylinder_result.txt", "w", encoding="utf-8") as f:
        f.write("==== 2025国赛A题 问题5 圆柱严格全遮蔽 结果 ====\n")
        f.write("判据: 云团到导弹-圆柱表面采样点视线段距离均 <= 10 m\n")
        f.write(f"三导弹有效遮蔽时长之和 = {total_sum:.6f} s\n")
        for m in MIS_NAMES:
            f.write(f"{m}: {missile_totals[m]:.6f} s  segs={missile_details[m]['segments']}\n")
        for b in all_bombs:
            f.write(
                f"{b['uav']} bomb{b['bomb_id']} -> {b['missile']}: "
                f"heading={b['heading_deg']:.6f} v={b['v']:.4f} "
                f"t_drop={b['t_drop']:.4f} tau={b['tau']:.4f} dur={b['duration']:.4f}\n"
            )
            f.write(f"  P_drop={b['P_drop']}\n")
            f.write(f"  P_det={b['P_det']}\n")
        f.write(f"device={device_info()}\n")
        f.write(f"elapsed={elapsed:.2f}s\n")

    summary = (
        f"M1={missile_totals['M1']:.4f}s, M2={missile_totals['M2']:.4f}s, "
        f"M3={missile_totals['M3']:.4f}s, sum={total_sum:.4f}s （圆柱严格全遮蔽）"
    )
    xlsx_path = OUT / "result3.xlsx"
    write_result3_xlsx(all_bombs, xlsx_path, summary=summary)
    print(f"已写 {display_path(xlsx_path)} 与 附件/result3.xlsx")
    print(f"总耗时 {elapsed:.1f}s")


if __name__ == "__main__":
    main()

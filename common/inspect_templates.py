# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
base = REPO_ROOT / "附件"
for name in ["result1.xlsx", "result2.xlsx", "result3.xlsx"]:
    p = base / name
    wb = load_workbook(p)
    ws = wb.active
    print("====", name, "sheet", ws.title, "dims", ws.dimensions)
    for i, row in enumerate(ws.iter_rows(max_row=12, max_col=16, values_only=True), 1):
        print(i, row)
    print()

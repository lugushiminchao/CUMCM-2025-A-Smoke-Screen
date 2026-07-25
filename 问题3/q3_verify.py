# -*- coding: utf-8 -*-
"""
问题3 结果验证：用问题1/问题2 同一套几何模型逐一校验

校验项：
1. 问题1 固定策略复现（基准）
2. 问题2 最优单弹复现
3. 问题3 三弹各自：运动学、投放间隔、速度边界、单弹遮蔽时长
4. 三弹并集时长（与单弹模型一致的 LOS 判据）
5. 与 q3_result.json / result1.xlsx 数值一致性
"""
from __future__ import annotations

import json
import math
import sys
from pathlib import Path

import numpy as np

# 直接复用问题2 的“第一问模型”
ROOT3 = Path(__file__).resolve().parent
REPO_ROOT = ROOT3.parent
Q2_DIR = REPO_ROOT / "问题2"
sys.path.insert(0, str(Q2_DIR))
from q2_optimize import (  # noqa: E402
    FY0,
    G,
    M0,
    R,
    T_HIT,
    T_LIFE,
    TGT,
    V_M,
    V_MISSILE,
    V_SINK,
    kinematics,
    los_dist,
    shield_duration,
)

OUT = ROOT3 / "结果"
RESULT_JSON = OUT / "q3_result.json"
RESULT_XLSX = OUT / "result1.xlsx"
LOG = OUT / "q3_verify_log.txt"

lines = []


def log(*a):
    s = " ".join(str(x) for x in a)
    print(s)
    lines.append(s)


def shield_segments(p_det, t_det, n=300000, bisect=50):
    """与问题1/2 相同判据，返回总时长与所有有效区间。"""
    t0 = float(t_det)
    t1 = min(float(t_det) + T_LIFE, T_HIT)
    if t1 <= t0:
        return 0.0, []
    ts = np.linspace(t0, t1, n)
    d, s = los_dist(ts, p_det, t_det)
    ok = (d <= R) & (s >= 0.0) & (s <= 1.0)
    if not np.any(ok):
        return 0.0, []

    jump = np.where(np.diff(ok.astype(np.int8)) != 0)[0]
    edges = []
    for j in jump:
        lo, hi = float(ts[j]), float(ts[j + 1])
        left_ok = bool(ok[j])
        for _ in range(bisect):
            mid = 0.5 * (lo + hi)
            dm, sm = los_dist(np.array([mid]), p_det, t_det)
            mid_ok = bool((dm[0] <= R) and (0.0 <= sm[0] <= 1.0))
            if mid_ok == left_ok:
                lo = mid
            else:
                hi = mid
        edges.append(0.5 * (lo + hi))

    segs = []
    state = bool(ok[0])
    cur = t0
    for e in edges:
        if state:
            segs.append((cur, e))
            state = False
        else:
            cur = e
            state = True
    if state:
        segs.append((cur, t1))
    dur = float(sum(b - a for a, b in segs))
    return dur, segs


def union_from_masks(bombs, n=500000, bisect=50):
    """bombs: list of (p_det, t_det)；并集时长 + 分段。"""
    t_dets = [float(td) for _, td in bombs]
    t0 = max(0.0, min(t_dets) - 0.1)
    t1 = min(T_HIT, max(t_dets) + T_LIFE)
    ts = np.linspace(t0, t1, n)
    any_ok = np.zeros(n, dtype=bool)
    per_ok = []
    for p_det, t_det in bombs:
        d, s = los_dist(ts, p_det, t_det)
        alive = (ts >= t_det) & (ts <= t_det + T_LIFE)
        ok = alive & (d <= R) & (s >= 0.0) & (s <= 1.0)
        per_ok.append(ok)
        any_ok |= ok

    # 单弹网格时长（粗）
    dt = ts[1] - ts[0]
    per_grid = [float(np.sum(ok) * dt) for ok in per_ok]

    if not np.any(any_ok):
        return 0.0, [], per_grid

    jump = np.where(np.diff(any_ok.astype(np.int8)) != 0)[0]
    edges = []
    for j in jump:
        lo, hi = float(ts[j]), float(ts[j + 1])
        left_ok = bool(any_ok[j])
        for _ in range(bisect):
            mid = 0.5 * (lo + hi)
            # 任一一弹在 mid 遮蔽
            mid_ok = False
            for p_det, t_det in bombs:
                if mid < t_det or mid > t_det + T_LIFE:
                    continue
                dm, sm = los_dist(np.array([mid]), p_det, t_det)
                if (dm[0] <= R) and (0.0 <= sm[0] <= 1.0):
                    mid_ok = True
                    break
            if mid_ok == left_ok:
                lo = mid
            else:
                hi = mid
        edges.append(0.5 * (lo + hi))

    segs = []
    state = bool(any_ok[0])
    cur = t0
    for e in edges:
        if state:
            segs.append((cur, e))
            state = False
        else:
            cur = e
            state = True
    if state:
        segs.append((cur, t1))
    dur = float(sum(b - a for a, b in segs))
    return dur, segs, per_grid


def approx_eq(a, b, atol=1e-3, rtol=1e-4):
    a, b = float(a), float(b)
    return abs(a - b) <= atol + rtol * max(abs(a), abs(b), 1.0)


def main():
    log("=" * 60)
    log("问题3 结果验证（复用问题1/2 几何模型）")
    log("=" * 60)
    log(f"T_HIT = {T_HIT:.9f} s")
    log(f"TGT  = {TGT.tolist()}")
    log(f"R    = {R} m, V_SINK={V_SINK}, T_LIFE={T_LIFE}")

    n_fail = 0

    # ---------- 1. 问题1 基准 ----------
    log("\n[1] 问题1 固定策略复现")
    # FY1: v=120 朝假目标（-x），t_drop=1.5, tau=3.6
    v_fy1 = np.array([-120.0, 0.0, 0.0])
    p_drop1 = FY0 + v_fy1 * 1.5
    t_det1 = 1.5 + 3.6
    p_det1 = p_drop1 + np.array([v_fy1[0] * 3.6, v_fy1[1] * 3.6, -0.5 * G * 3.6**2])
    d1, segs1 = shield_segments(p_det1, t_det1, n=400000, bisect=55)
    log(f"  P_drop = {p_drop1.tolist()}")
    log(f"  P_det  = {np.round(p_det1, 6).tolist()}")
    log(f"  t_det  = {t_det1}")
    log(f"  时长   = {d1:.6f} s  segs={[(round(a,6),round(b,6)) for a,b in segs1]}")
    # 问题1 已知约 1.4055
    if 1.40 <= d1 <= 1.42:
        log("  PASS: 与问题1量级一致 (~1.4055 s)")
    else:
        log(f"  FAIL: 问题1 复现异常 d={d1}")
        n_fail += 1

    # ---------- 2. 问题2 最优 ----------
    log("\n[2] 问题2 最优单弹复现")
    th2, v2, td2, tau2 = 3.0880757565, 71.8890217683, 0.0, 2.5032397513
    v_fy2, p_drop2, t_det2, p_det2 = kinematics(th2, v2, td2, tau2)
    d2, segs2 = shield_segments(p_det2, t_det2, n=400000, bisect=55)
    log(f"  P_det = {np.round(p_det2, 6).tolist()}")
    log(f"  t_det = {t_det2:.10f}")
    log(f"  时长  = {d2:.6f} s  segs={[(round(a,6),round(b,6)) for a,b in segs2]}")
    if abs(d2 - 4.736887) < 0.01:
        log("  PASS: 与问题2结果 4.736887 s 一致")
    else:
        log(f"  FAIL: 问题2 复现 d={d2} vs 4.736887")
        n_fail += 1

    # ---------- 3. 加载问题3 ----------
    log("\n[3] 加载问题3结果并校验约束/运动学")
    data = json.load(open(RESULT_JSON, encoding="utf-8"))
    info = data["info"]
    theta = float(info["theta"])
    v = float(info["v"])
    heading = float(info["heading_deg"])
    bombs_info = info["bombs"]

    log(f"  heading={heading:.6f} deg, theta={theta:.10f}, v={v:.10f}")

    # 速度边界
    if 70 - 1e-6 <= v <= 140 + 1e-6:
        log("  PASS: 速度在 [70,140]")
    else:
        log(f"  FAIL: 速度越界 v={v}")
        n_fail += 1

    # 航向一致性
    if abs((math.degrees(theta) % 360) - (heading % 360)) < 1e-4:
        log("  PASS: heading 与 theta 一致")
    else:
        log("  FAIL: heading/theta 不一致")
        n_fail += 1

    bombs_model = []  # (p_det, t_det, p_drop, t_drop, tau, dur_q1model, segs)
    t_drops = []
    for i, b in enumerate(bombs_info):
        t_drop = float(b["t_drop"])
        tau = float(b["tau"])
        t_drops.append(t_drop)
        v_fy, p_drop_calc, t_det_calc, p_det_calc = kinematics(theta, v, t_drop, tau)

        p_drop_json = np.array(b["P_drop"], dtype=float)
        p_det_json = np.array(b["P_det"], dtype=float)
        t_det_json = float(b["t_det"])

        err_drop = float(np.linalg.norm(p_drop_calc - p_drop_json))
        err_det = float(np.linalg.norm(p_det_calc - p_det_json))
        err_tdet = abs(t_det_calc - t_det_json)

        log(f"\n  --- 弹 {i+1} 运动学 ---")
        log(f"  t_drop={t_drop:.10f}, tau={tau:.10f}, t_det={t_det_calc:.10f}")
        log(f"  P_drop calc={np.round(p_drop_calc,6).tolist()}")
        log(f"  P_det  calc={np.round(p_det_calc,6).tolist()}")
        log(f"  err_drop={err_drop:.3e}, err_det={err_det:.3e}, err_tdet={err_tdet:.3e}")

        if err_drop < 1e-4 and err_det < 1e-4 and err_tdet < 1e-8:
            log("  PASS: JSON 与运动学一致")
        else:
            log("  FAIL: 运动学不一致")
            n_fail += 1

        # 等高：投放 z=1800
        if abs(p_drop_calc[2] - 1800.0) < 1e-6:
            log("  PASS: 投放高度 1800 m")
        else:
            log("  FAIL: 投放高度异常")
            n_fail += 1

        # 起爆高度：自由落体
        z_expect = 1800.0 - 0.5 * G * tau * tau
        if abs(p_det_calc[2] - z_expect) < 1e-6:
            log("  PASS: 起爆高度 = 1800 - 0.5*g*tau^2")
        else:
            log("  FAIL: 起爆高度公式")
            n_fail += 1

        # 单弹遮蔽（问题1模型）
        dur_i, segs_i = shield_segments(p_det_calc, t_det_calc, n=400000, bisect=55)
        log(f"  单弹遮蔽时长(Q1模型) = {dur_i:.6f} s")
        log(f"  单弹区间 = {[(round(a,6), round(b,6), round(b-a,6)) for a,b in segs_i]}")
        reported = float(data["per_bomb"][i])
        if approx_eq(dur_i, reported, atol=0.02, rtol=0.005):
            log(f"  PASS: 与报告单弹时长 {reported:.6f} 一致")
        else:
            log(f"  WARN/FAIL: 报告 {reported:.6f} vs 复算 {dur_i:.6f}")
            if abs(dur_i - reported) > 0.05:
                n_fail += 1

        bombs_model.append(
            {
                "p_det": p_det_calc,
                "t_det": t_det_calc,
                "p_drop": p_drop_calc,
                "t_drop": t_drop,
                "tau": tau,
                "dur": dur_i,
                "segs": segs_i,
            }
        )

    # 投放间隔
    log("\n[4] 投放间隔约束")
    gaps = [t_drops[1] - t_drops[0], t_drops[2] - t_drops[1]]
    log(f"  gaps = {gaps}")
    if gaps[0] >= 1.0 - 1e-9 and gaps[1] >= 1.0 - 1e-9 and t_drops[0] >= -1e-9:
        log("  PASS: 投放时刻非负且间隔 >= 1 s")
    else:
        log("  FAIL: 投放间隔/时刻约束")
        n_fail += 1
    # 顺序
    if t_drops[0] <= t_drops[1] <= t_drops[2]:
        log("  PASS: 投放顺序 t1<=t2<=t3")
    else:
        log("  FAIL: 投放顺序")
        n_fail += 1

    # ---------- 5. 并集 ----------
    log("\n[5] 三弹并集（问题1 LOS 模型）")
    bomb_pairs = [(bm["p_det"], bm["t_det"]) for bm in bombs_model]
    union_dur, union_segs, per_grid = union_from_masks(bomb_pairs, n=600000, bisect=55)
    log(f"  并集时长 = {union_dur:.6f} s")
    log(f"  并集分段 = {[(round(a,6), round(b,6), round(b-a,6)) for a,b in union_segs]}")
    log(f"  网格单弹时长 ≈ {np.round(per_grid, 6).tolist()}")
    sum_per = sum(bm["dur"] for bm in bombs_model)
    log(f"  单弹时长和 = {sum_per:.6f} s, 重叠估计 = {sum_per - union_dur:.6f} s")
    reported_u = float(data["total_union"])
    if approx_eq(union_dur, reported_u, atol=0.02, rtol=0.005):
        log(f"  PASS: 与报告并集 {reported_u:.6f} 一致")
    else:
        log(f"  FAIL: 报告并集 {reported_u:.6f} vs 复算 {union_dur:.6f}")
        n_fail += 1

    # 并集应 >= 任一单弹，且 <= 单弹和
    max_single = max(bm["dur"] for bm in bombs_model)
    if union_dur + 1e-6 >= max_single and union_dur <= sum_per + 1e-3:
        log("  PASS: max(单弹) <= 并集 <= sum(单弹)")
    else:
        log("  FAIL: 并集与单弹关系不合理")
        n_fail += 1

    # 应明显优于单弹 Q2
    if union_dur > d2 + 0.5:
        log(f"  PASS: 并集 {union_dur:.3f} > Q2单弹 {d2:.3f}")
    else:
        log("  WARN: 并集未明显超过 Q2")

    # ---------- 6. 抽样时刻点检 ----------
    log("\n[6] 关键时刻 LOS 点检")
    # 在并集中点、各弹预计有效段中点
    check_ts = []
    for a, b in union_segs:
        check_ts.append(0.5 * (a + b))
        check_ts.append(a + 0.05)
        check_ts.append(b - 0.05)
    for bm in bombs_model:
        for a, b in bm["segs"]:
            check_ts.append(0.5 * (a + b))
    # 并集外一点
    if union_segs:
        check_ts.append(union_segs[0][0] - 0.5)
        check_ts.append(union_segs[-1][1] + 0.5)

    for t in sorted(set(round(tt, 6) for tt in check_ts)):
        if t < 0 or t > T_HIT:
            continue
        flags = []
        for i, bm in enumerate(bombs_model):
            if t < bm["t_det"] or t > bm["t_det"] + T_LIFE:
                flags.append(False)
                continue
            d, s = los_dist(np.array([t]), bm["p_det"], bm["t_det"])
            flags.append(bool((d[0] <= R) and (0.0 <= s[0] <= 1.0)))
        anyf = any(flags)
        # 是否应在并集内
        in_union = any(a - 1e-6 <= t <= b + 1e-6 for a, b in union_segs)
        status = "OK" if anyf == in_union or abs(t - union_segs[0][0]) < 0.1 or abs(t - union_segs[-1][1]) < 0.1 else "CHECK"
        # 边界附近允许
        near_edge = any(abs(t - a) < 0.15 or abs(t - b) < 0.15 for a, b in union_segs)
        if anyf != in_union and not near_edge:
            status = "FAIL"
            n_fail += 1
        elif anyf != in_union and near_edge:
            status = "EDGE"
        log(f"  t={t:10.6f} bombs={flags} any={anyf} in_union_seg={in_union} [{status}]")

    # ---------- 7. Excel 一致性 ----------
    log("\n[7] result1.xlsx 一致性")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(RESULT_XLSX)
        ws = wb.active
        for i, bm in enumerate(bombs_model):
            r = 2 + i
            h_x = float(ws.cell(r, 1).value)
            v_x = float(ws.cell(r, 2).value)
            id_x = int(ws.cell(r, 3).value)
            drop_x = np.array([float(ws.cell(r, 4 + j).value) for j in range(3)])
            det_x = np.array([float(ws.cell(r, 7 + j).value) for j in range(3)])
            dur_x = float(ws.cell(r, 10).value)
            ok_h = approx_eq(h_x, heading, atol=1e-4)
            ok_v = approx_eq(v_x, v, atol=1e-4)
            ok_id = id_x == i + 1
            ok_drop = np.linalg.norm(drop_x - bm["p_drop"]) < 0.01
            ok_det = np.linalg.norm(det_x - bm["p_det"]) < 0.01
            ok_dur = approx_eq(dur_x, bm["dur"], atol=0.02)
            all_ok = ok_h and ok_v and ok_id and ok_drop and ok_det and ok_dur
            log(
                f"  行{r}: id={id_x} h={h_x} v={v_x} dur={dur_x} "
                f"drop_err={np.linalg.norm(drop_x-bm['p_drop']):.3e} "
                f"det_err={np.linalg.norm(det_x-bm['p_det']):.3e} "
                f"{'PASS' if all_ok else 'FAIL'}"
            )
            if not all_ok:
                n_fail += 1
        note = ws.cell(5, 1).value
        log(f"  备注行: {note}")
    except Exception as e:
        log(f"  FAIL: 读取 Excel 异常 {e}")
        n_fail += 1

    # ---------- 8. 水平速度一致性（等高直线） ----------
    log("\n[8] 水平速度/直线飞行一致性")
    vx = v * math.cos(theta)
    vy = v * math.sin(theta)
    for i, bm in enumerate(bombs_model):
        # 从 FY0 到 drop 的平均速度
        dt = bm["t_drop"]
        if dt > 1e-9:
            v_avg = (bm["p_drop"] - FY0) / dt
            err_v = abs(v_avg[0] - vx) + abs(v_avg[1] - vy) + abs(v_avg[2])
            log(f"  弹{i+1} 投放前均速 err={err_v:.3e} (期望 vx,vy,0=[{vx:.6f},{vy:.6f},0])")
            if err_v > 1e-4:
                n_fail += 1
                log("  FAIL")
            else:
                log("  PASS")
        else:
            err = float(np.linalg.norm(bm["p_drop"] - FY0))
            log(f"  弹{i+1} t_drop≈0, |P_drop-FY0|={err:.3e}")
            if err > 1.0:
                n_fail += 1

    # ---------- 总结 ----------
    log("\n" + "=" * 60)
    log(f"验证完成: FAIL 计数 = {n_fail}")
    if n_fail == 0:
        log("结论: 全部通过 — 问题3结果与问题1/2 模型一致")
    else:
        log("结论: 存在未通过项，请检查上方 FAIL")
    log("=" * 60)

    # 汇总表
    log("\n【汇总】")
    log(f"  Q1 复现时长     = {d1:.6f} s")
    log(f"  Q2 复现时长     = {d2:.6f} s")
    for i, bm in enumerate(bombs_model):
        log(f"  Q3 弹{i+1} 时长    = {bm['dur']:.6f} s  t_det={bm['t_det']:.4f}")
    log(f"  Q3 并集时长     = {union_dur:.6f} s")
    log(f"  报告并集时长    = {reported_u:.6f} s")
    log(f"  相对 Q2 提升    = {union_dur - d2:.6f} s")

    OUT.mkdir(exist_ok=True)
    with open(LOG, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    log(f"\n日志已写: {LOG.relative_to(REPO_ROOT).as_posix()}")
    return n_fail


if __name__ == "__main__":
    code = main()
    raise SystemExit(0 if code == 0 else 1)

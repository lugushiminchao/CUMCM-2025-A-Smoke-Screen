# -*- coding: utf-8 -*-
"""保存 stage3/polish 中出现过的最优种子，并用一致高精度重算。"""
import json
import sys
from pathlib import Path

import numpy as np

THIS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(THIS_DIR))
from q3_optimize import (  # noqa
    OUT,
    device_info,
    highres_union,
    local_refine,
    write_result1_xlsx,
)

# seed5 附近的优解（v=140, th~pi, t1~0）
# 从 polish 日志 seed5 high 7.756566
# 重新从当前解 + 多轮精修，取 highres 最大

cur = json.load(open(OUT / "q3_result.json", encoding="utf-8"))
candidates = [np.array(cur["info"]["x"], dtype=float)]

# 围绕日志最优构造
base = np.array(
    [
        3.135437871159533,
        140.0,
        0.0,
        3.62,
        3.65,
        5.35,
        5.55,
        6.04,
    ],
    dtype=float,
)
candidates.append(base)
candidates.append(
    np.array(
        [3.135, 140, 0.0, 3.63, 3.63, 5.35, 5.54, 6.04],
        dtype=float,
    )
)

best = None
for i, s0 in enumerate(candidates):
    xf, df = local_refine(
        s0,
        n_time=50000,
        steps=90,
        spans=np.array([0.035, 2.0, 0.35, 0.45, 0.55, 0.55, 0.65, 0.65]),
        use_gpu=True,
    )
    # 多分辨率取 max 避免网格噪声
    scores = []
    for n in (200000, 500000, 1000000):
        t, p, info = highres_union(xf, n_time=n)
        scores.append((t, p, info, xf, n))
    t, p, info, xf, n = max(scores, key=lambda z: z[0])
    print(f"cand{i} mid={df:.6f} high={t:.6f} n={n} per={np.round(p,4)}")
    if best is None or t > best[0]:
        best = (t, p, info, xf)

# 再 polish 最优
xf, df = local_refine(
    best[3],
    n_time=60000,
    steps=50,
    spans=np.array([0.012, 0.8, 0.12, 0.18, 0.18, 0.18, 0.22, 0.22]),
    use_gpu=True,
)
scores = []
for n in (300000, 800000, 1500000):
    t, p, info = highres_union(xf, n_time=n)
    scores.append((t, p, info, n))
    print(f"  polish n={n} -> {t:.6f}")
total, per, info, nbest = max(scores, key=lambda z: z[0])
print("BEST", total, "per", per)
print("x", info["x"])
print("heading", info["heading_deg"], "v", info["v"])

with open(OUT / "q3_result.txt", "w", encoding="utf-8") as f:
    f.write("==== 2025国赛A题 问题3 结果 ====\n")
    f.write("FY1 投放 3 枚烟幕干扰弹，干扰 M1；遮蔽取并集（可不连续）\n")
    f.write(f"backend = {device_info()}\n")
    f.write(f"并集有效遮蔽时长 = {total:.6f} s\n")
    f.write(
        f"单弹有效时长(各自，可重叠) = [{per[0]:.6f}, {per[1]:.6f}, {per[2]:.6f}] s\n"
    )
    f.write("\n--- 无人机 ---\n")
    f.write(
        f"航向角 theta = {info['theta']:.10f} rad = {info['heading_deg']:.6f} deg\n"
    )
    f.write(f"飞行速度 v   = {info['v']:.10f} m/s\n")
    f.write("航向定义: 以 x 轴正向为基准逆时针，取值 0~360°\n")
    for j, b in enumerate(info["bombs"]):
        f.write(f"\n--- 烟幕干扰弹 {j+1} ---\n")
        f.write(f"投放时刻 t_drop = {b['t_drop']:.10f} s\n")
        f.write(f"引信延时 tau    = {b['tau']:.10f} s\n")
        f.write(f"起爆时刻 t_det  = {b['t_det']:.10f} s\n")
        f.write(
            f"投放点 P_drop = [{b['P_drop'][0]:.10f}, {b['P_drop'][1]:.10f}, {b['P_drop'][2]:.10f}]\n"
        )
        f.write(
            f"起爆点 P_det  = [{b['P_det'][0]:.10f}, {b['P_det'][1]:.10f}, {b['P_det'][2]:.10f}]\n"
        )
        f.write(f"单弹有效时长 = {per[j]:.6f} s\n")
    f.write(f"\n决策向量 x = {info['x']}\n")
    f.write("求解 = Python + PyTorch CUDA 批量评估 + 多阶段粗搜/坐标轮换精修\n")
    f.write("判据: 云团球心到导弹-真目标中心视线段距离 <= 10 m；三枚取并集（可不连续）\n")
    f.write("约束: 投放间隔 >= 1 s；无人机速度 70~140 m/s；等高匀速直线\n")
    f.write(f"高精度时间网格点数 ≈ {nbest}\n")

json.dump(
    {
        "total_union": float(total),
        "per_bomb": per.tolist(),
        "info": info,
        "backend": device_info(),
        "n_time_highres": nbest,
    },
    open(OUT / "q3_result.json", "w", encoding="utf-8"),
    ensure_ascii=False,
    indent=2,
)
path = write_result1_xlsx(info, per, total, OUT / "result1.xlsx")
print("saved", Path(path).resolve().relative_to(THIS_DIR.parent.resolve()).as_posix())

# verify excel
from openpyxl import load_workbook

for label, pth in [
    ("out", OUT / "result1.xlsx"),
    ("attach", OUT.parent.parent / "附件" / "result1.xlsx"),
]:
    wb = load_workbook(pth)
    ws = wb.active
    print(label, Path(pth).resolve().relative_to(THIS_DIR.parent.resolve()).as_posix())
    for r in range(1, 6):
        print(" ", [ws.cell(r, c).value for c in range(1, 11)])

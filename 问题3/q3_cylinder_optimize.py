# -*- coding: utf-8 -*-
"""
问题3：圆柱严格全遮蔽下，FY1 投放 3 枚烟幕干扰弹，最大化并集有效遮蔽时长
"""
from __future__ import annotations

import json
import shutil
import sys
import time
from pathlib import Path

import numpy as np

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
from common.smoke_geom import (  # noqa: E402
    DROP_GAP,
    FY0,
    G,
    PTS_FAST,
    PTS_FINE,
    PTS_OPT,
    T_HIT,
    T_LIFE,
    device_info,
    kinematics,
    kinematics_batch,
    shield_duration,
    union_duration_batch,
    union_duration_single_path,
)

OUT = ROOT / "结果"

def display_path(path):
    return Path(path).resolve().relative_to(ROOT.parent.resolve()).as_posix()

OUT.mkdir(exist_ok=True)
ATTACH = ROOT.parent / "附件"

MODE = "cylinder"
PTS_COARSE = PTS_FAST
PTS_MID = PTS_OPT
PTS_HI = PTS_FINE


def fix_times(X):
    X = np.asarray(X, dtype=np.float64).copy()
    if X.ndim == 1:
        X = X[None, :]
        single = True
    else:
        single = False
    t1 = np.maximum(X[:, 2], 0.0)
    t2 = np.maximum(X[:, 4], t1 + DROP_GAP)
    t3 = np.maximum(X[:, 6], t2 + DROP_GAP)
    X[:, 2], X[:, 4], X[:, 6] = t1, t2, t3
    X[:, 3] = np.maximum(X[:, 3], 0.2)
    X[:, 5] = np.maximum(X[:, 5], 0.2)
    X[:, 7] = np.maximum(X[:, 7], 0.2)
    X[:, 1] = np.clip(X[:, 1], 70.0, 140.0)
    X[:, 0] = np.mod(X[:, 0], 2 * np.pi)
    return X[0] if single else X


def eval_strategies(X, n_time=5000, use_gpu=True, pts=None):
    X = fix_times(np.asarray(X, dtype=np.float64))
    if X.ndim == 1:
        X = X[None, :]
    if pts is None:
        pts = PTS_COARSE
    theta, v = X[:, 0], X[:, 1]
    t1, tau1 = X[:, 2], X[:, 3]
    t2, tau2 = X[:, 4], X[:, 5]
    t3, tau3 = X[:, 6], X[:, 7]
    dets = []
    tds = []
    for ti, taui in ((t1, tau1), (t2, tau2), (t3, tau3)):
        _, pdet, td = kinematics_batch(theta, v, ti, taui)
        dets.append(pdet)
        tds.append(td)
    P_det = np.stack(dets, axis=1)
    t_det = np.stack(tds, axis=1)
    dur, per = union_duration_batch(
        P_det, t_det, mode=MODE, pts=pts, n_time=n_time, use_gpu=use_gpu
    )
    return dur, per, P_det, t_det


def pack_strategy(x):
    x = fix_times(np.asarray(x, dtype=float))
    bombs = []
    for i in range(3):
        td = float(x[2 + 2 * i])
        tau = float(x[3 + 2 * i])
        v_fy, p_drop, t_det, p_det = kinematics(float(x[0]), float(x[1]), td, tau)
        bombs.append(
            {
                "t_drop": td,
                "tau": tau,
                "t_det": float(t_det),
                "P_drop": p_drop.tolist(),
                "P_det": p_det.tolist(),
            }
        )
    return {
        "theta": float(x[0]),
        "heading_deg": float(np.degrees(x[0]) % 360.0),
        "v": float(x[1]),
        "bombs": bombs,
        "x": x.tolist(),
        "v_fy": [
            float(x[1] * np.cos(x[0])),
            float(x[1] * np.sin(x[0])),
            0.0,
        ],
    }


def highres_union(x, pts=None, n=200000):
    if pts is None:
        pts = PTS_HI
    info = pack_strategy(x)
    bombs = [(np.array(b["P_det"]), b["t_det"]) for b in info["bombs"]]
    total, per, segs = union_duration_single_path(
        bombs, mode=MODE, pts=pts, n=n, refine=True
    )
    return total, per, segs, info


def local_refine(x0, n_time=5000, steps=50, spans=None, use_gpu=True, pts=None):
    if spans is None:
        spans = np.array([0.05, 5.0, 0.6, 0.6, 0.6, 0.6, 0.6, 0.6])
    if pts is None:
        pts = PTS_MID
    lb = np.array([0.0, 70.0, 0.0, 0.3, 1.0, 0.3, 2.0, 0.3])
    ub = np.array([2 * np.pi, 140.0, 20.0, 16.0, 25.0, 16.0, 30.0, 16.0])
    x = fix_times(np.array(x0, dtype=float))
    best = float(eval_strategies(x[None, :], n_time=n_time, use_gpu=use_gpu, pts=pts)[0][0])
    span = spans.copy()
    rng = np.random.default_rng(2025)
    for it in range(steps):
        improved = False
        for d in range(8):
            for sgn in (-1.0, 1.0):
                trial = x.copy()
                trial[d] += sgn * span[d]
                trial = fix_times(np.clip(trial, lb, ub))
                val = float(
                    eval_strategies(trial[None, :], n_time=n_time, use_gpu=use_gpu, pts=pts)[0][0]
                )
                if val > best + 1e-9:
                    best = val
                    x = trial
                    improved = True
        noise = (rng.random((40, 8)) - 0.5) * 2.0 * span[None, :]
        trials = fix_times(np.clip(x[None, :] + noise, lb, ub))
        vals = eval_strategies(trials, n_time=n_time, use_gpu=use_gpu, pts=pts)[0]
        j = int(np.argmax(vals))
        if vals[j] > best + 1e-9:
            best = float(vals[j])
            x = trials[j].copy()
            improved = True
        if improved:
            span *= 0.90
        else:
            span = np.minimum(span * 1.06, spans * 1.25)
        if (it + 1) % 12 == 0:
            print(f"    refine it={it+1} best={best:.4f}")
    return x, best


def write_result1_xlsx(info, per, total, path: Path):
    # 优先套模板
    template = ATTACH / "result1.xlsx"
    if load_workbook is not None and template.exists():
        wb = load_workbook(template)
        ws = wb.active
    else:
        if Workbook is None:
            raise RuntimeError("openpyxl required")
        wb = Workbook()
        ws = wb.active
        headers = [
            "无人机运动方向",
            "无人机运动速度 (m/s)",
            "烟幕干扰弹编号",
            "烟幕干扰弹投放点的x坐标 (m)",
            "烟幕干扰弹投放点的y坐标 (m)",
            "烟幕干扰弹投放点的z坐标 (m)",
            "烟幕干扰弹起爆点的x坐标 (m)",
            "烟幕干扰弹起爆点的y坐标 (m)",
            "烟幕干扰弹起爆点的z坐标 (m)",
            "有效干扰时长 (s)",
        ]
        for j, h in enumerate(headers, 1):
            ws.cell(1, j, h)

    heading = info["heading_deg"]
    v = info["v"]
    for i, b in enumerate(info["bombs"]):
        row = i + 2
        ws.cell(row, 1, float(heading))
        ws.cell(row, 2, float(v))
        ws.cell(row, 3, i + 1)
        for j in range(3):
            ws.cell(row, 4 + j, float(b["P_drop"][j]))
            ws.cell(row, 7 + j, float(b["P_det"][j]))
        ws.cell(row, 10, float(per[i]))
    ws.cell(5, 1, f"三枚并集有效遮蔽总时长 = {total:.6f} s （圆柱严格全遮蔽）")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    # 同步附件
    try:
        shutil.copy2(path, ATTACH / "result1.xlsx")
    except Exception as e:
        print("copy to 附件 failed:", e)


def build_candidates(rng):
    rows = []
    # warm starts from point-model Q2/Q3 and cylinder Q2 optimum
    q2_th, q2_v, q2_tau = 3.0825785063, 70.0, 2.4841194504
    q3_pt = np.array(
        [
            3.135440402528894,
            139.99827578912254,
            0.0029830233034361686,
            3.611130468416467,
            3.7024597386105236,
            5.337486706836162,
            5.56948399858038,
            6.040529895407159,
        ]
    )
    rows.append(q3_pt)
    rows.append([np.pi, 120, 1.5, 3.6, 2.5, 3.6, 3.5, 3.6])
    rows.append([np.pi, 140, 0.0, 2.5, 1.0, 3.0, 2.0, 3.5])
    rows.append([np.pi, 140, 0.0, 3.5, 1.5, 4.5, 3.0, 5.5])
    rows.append([q2_th, q2_v, 0.0, q2_tau, 1.0, q2_tau, 2.0, q2_tau])
    rows.append([q2_th, q2_v, 0.0, q2_tau, 1.2, 3.0, 2.5, 3.5])
    rows.append([q2_th, 80.0, 0.0, q2_tau, 1.5, 3.5, 3.0, 4.5])
    rows.append([np.pi, 100, 0.0, 2.5, 1.0, 3.5, 2.0, 4.5])
    rows.append([np.pi, 90, 0.0, 3.0, 1.2, 4.0, 2.5, 5.0])
    rows.append([np.pi, 130, 0.0, 3.0, 1.5, 4.0, 3.0, 5.0])
    rows.append([np.pi, 70, 0.0, 3.5, 1.0, 4.0, 2.0, 4.5])

    ths = np.pi + np.linspace(-0.30, 0.30, 11)
    vs = np.array([70, 90, 110, 120, 130, 140], dtype=float)
    drop_sets = [
        (0.0, 1.2, 2.5),
        (0.0, 1.5, 3.0),
        (0.0, 2.0, 4.0),
        (0.2, 1.5, 3.0),
        (0.5, 2.0, 4.0),
        (0.0, 1.0, 3.5),
        (0.0, 2.5, 5.0),
        (1.0, 2.5, 4.5),
        (0.0, 3.5, 5.5),
    ]
    tau_sets = [
        (2.5, 3.5, 4.5),
        (3.0, 4.0, 5.0),
        (3.5, 4.5, 5.5),
        (2.0, 3.0, 4.0),
        (3.5, 5.0, 6.0),
        (2.5, 4.0, 5.5),
        (4.0, 5.0, 6.0),
        (3.0, 5.0, 6.5),
    ]
    for th in ths:
        for v in vs:
            for drops in drop_sets:
                for taus in tau_sets:
                    rows.append([th, v, drops[0], taus[0], drops[1], taus[1], drops[2], taus[2]])

    # random
    n_rand = 4000
    Xr = np.column_stack(
        [
            (np.pi + (rng.random(n_rand) - 0.5) * 0.7) % (2 * np.pi),
            70 + rng.random(n_rand) * 70,
            rng.random(n_rand) * 8,
            0.5 + rng.random(n_rand) * 8,
            rng.random(n_rand) * 10,
            0.5 + rng.random(n_rand) * 9,
            rng.random(n_rand) * 12,
            0.5 + rng.random(n_rand) * 10,
        ]
    )
    X = fix_times(np.vstack([np.asarray(rows, dtype=float), Xr]))
    # unique rounded
    X = np.unique(np.round(X, 5), axis=0)
    return X


def main():
    t0 = time.perf_counter()
    use_gpu = True
    print("==== 问题3 圆柱严格全遮蔽 三弹并集优化 ====")
    print("device:", device_info())
    print(f"t_hit={T_HIT:.6f}")
    print(f"pts coarse/mid/hi = {PTS_COARSE.shape[0]}/{PTS_MID.shape[0]}/{PTS_HI.shape[0]}")

    # sanity Q1
    x_q1 = np.array([np.pi, 120, 1.5, 3.6, 2.5, 3.6, 3.5, 3.6])
    d0, p0, _, _ = eval_strategies(x_q1[None, :], n_time=6000, use_gpu=use_gpu, pts=PTS_COARSE)
    print(f"sanity Q1-like union coarse={d0[0]:.4f}")

    rng = np.random.default_rng(2025)
    X = build_candidates(rng)
    n_all = len(X)
    print(f"粗搜样本 {n_all}")

    t_c = time.perf_counter()
    chunk = 800
    durs = np.zeros(n_all)
    for i0 in range(0, n_all, chunk):
        i1 = min(n_all, i0 + chunk)
        durs[i0:i1], _, _, _ = eval_strategies(
            X[i0:i1], n_time=3500, use_gpu=use_gpu, pts=PTS_COARSE
        )
        print(
            f"  coarse {i1}/{n_all} best={durs[:i1].max():.4f} "
            f"({time.perf_counter()-t_c:.1f}s)"
        )

    order = np.argsort(-durs)
    Xs, ds = X[order], durs[order]
    print("Top-12 粗搜:")
    for k in range(min(12, n_all)):
        xk = Xs[k]
        print(
            f"  #{k+1:2d} {ds[k]:.4f} th={np.degrees(xk[0])%360:.2f} v={xk[1]:.1f} "
            f"t=[{xk[2]:.2f},{xk[4]:.2f},{xk[6]:.2f}] "
            f"tau=[{xk[3]:.2f},{xk[5]:.2f},{xk[7]:.2f}]"
        )

    n_top = 16
    print(f"---- 局部精修 Top-{n_top} ----")
    refined = []
    t_r = time.perf_counter()
    for k in range(min(n_top, n_all)):
        xf, df = local_refine(
            Xs[k],
            n_time=4500,
            steps=40,
            use_gpu=use_gpu,
            pts=PTS_MID,
            spans=np.array([0.05, 5.0, 0.7, 0.7, 0.7, 0.7, 0.7, 0.7]),
        )
        # mid-hi score
        d_hi, _, _, _ = eval_strategies(xf[None, :], n_time=10000, use_gpu=use_gpu, pts=PTS_MID)
        refined.append((float(d_hi[0]), xf, df))
        print(f"  {k+1}/{n_top}: {ds[k]:.4f} -> {df:.4f} / hi={d_hi[0]:.4f}")
    refined.sort(key=lambda t: -t[0])
    x_best = refined[0][1]
    print(f"精修最优 {refined[0][0]:.6f}  耗时 {time.perf_counter()-t_r:.1f}s")

    # 二次精修
    x_best, _ = local_refine(
        x_best,
        n_time=8000,
        steps=35,
        use_gpu=use_gpu,
        pts=PTS_MID,
        spans=np.array([0.025, 2.5, 0.35, 0.35, 0.35, 0.35, 0.35, 0.35]),
    )

    # 高精度
    total, per, segs, info = highres_union(x_best, pts=PTS_HI, n=220000)
    print("\n==== 问题3 圆柱最优 ====")
    print(f"并集时长 = {total:.6f} s")
    print(f"单弹 = {per}")
    print(f"分段 = {segs}")
    print(f"heading={info['heading_deg']:.6f} v={info['v']:.6f}")
    for i, b in enumerate(info["bombs"]):
        print(
            f"  bomb{i+1}: t_drop={b['t_drop']:.6f} tau={b['tau']:.6f} "
            f"t_det={b['t_det']:.6f} per={per[i]:.6f}"
        )

    elapsed = time.perf_counter() - t0
    result = {
        "total_union": total,
        "per_bomb": per,
        "segments": segs,
        "info": info,
        "mode": "cylinder_strict_full_cover",
        "device": device_info(),
        "elapsed_s": elapsed,
    }

    def conv(o):
        if isinstance(o, (np.floating,)):
            return float(o)
        if isinstance(o, (np.integer,)):
            return int(o)
        if isinstance(o, np.ndarray):
            return o.tolist()
        if isinstance(o, list):
            return [conv(x) for x in o]
        if isinstance(o, tuple):
            return [conv(x) for x in o]
        if isinstance(o, dict):
            return {k: conv(v) for k, v in o.items()}
        return o

    with open(OUT / "q3_cylinder_result.json", "w", encoding="utf-8") as f:
        json.dump(conv(result), f, ensure_ascii=False, indent=2)

    with open(OUT / "q3_cylinder_result.txt", "w", encoding="utf-8") as f:
        f.write("==== 2025国赛A题 问题3 圆柱严格全遮蔽 结果 ====\n")
        f.write("判据: 云团到导弹-圆柱表面采样点视线段距离均 <= 10 m（严格全遮蔽）\n")
        f.write(f"并集有效遮蔽时长 = {total:.6f} s\n")
        f.write(f"分段 = {segs}\n")
        f.write(f"单弹时长 = {per}\n")
        f.write(f"航向角 = {info['heading_deg']:.10f} deg\n")
        f.write(f"速度 = {info['v']:.10f} m/s\n")
        for i, b in enumerate(info["bombs"]):
            f.write(
                f"弹{i+1}: t_drop={b['t_drop']:.10f}, tau={b['tau']:.10f}, "
                f"t_det={b['t_det']:.10f}\n"
            )
            f.write(f"  P_drop={b['P_drop']}\n")
            f.write(f"  P_det={b['P_det']}\n")
            f.write(f"  单弹时长={per[i]:.6f}\n")
        f.write(f"device={device_info()}\n")
        f.write(f"elapsed={elapsed:.2f}s\n")

    xlsx_path = OUT / "result1.xlsx"
    write_result1_xlsx(info, per, total, xlsx_path)
    print(f"已写 {display_path(xlsx_path)} 与 附件/result1.xlsx")
    print(f"总耗时 {elapsed:.1f}s")


if __name__ == "__main__":
    main()

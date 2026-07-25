# -*- coding: utf-8 -*-
"""
2025国赛A题 问题3：FY1 投放 3 枚烟幕干扰弹干扰 M1
目标：最大化三枚烟幕有效遮蔽时间的并集（可不连续）

决策变量（8维）:
  theta, v, t1, tau1, t2, tau2, t3, tau3
约束:
  v ∈ [70,140], ti≥0, tau_i>0, t_{i+1} ≥ t_i + 1

加速:
  - 时间网格向量化
  - 批量评估大量候选（NumPy；若安装 torch+CUDA 则走 GPU）
  - 外层 ProcessPool 可选（默认关，避免小任务开销）
"""
from __future__ import annotations

import json
import math
import shutil
import time
from pathlib import Path

import numpy as np

try:
    import torch

    TORCH_OK = True
    CUDA_OK = torch.cuda.is_available()
except Exception:
    torch = None  # type: ignore
    TORCH_OK = False
    CUDA_OK = False

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "结果"

def display_path(path):
    return Path(path).resolve().relative_to(ROOT.parent.resolve()).as_posix()

OUT.mkdir(exist_ok=True)
ATTACH = ROOT.parent / "附件"

# ---------------- 场景常量 ----------------
G = 9.8
V_MISSILE = 300.0
V_SINK = 3.0
R = 10.0
R2 = R * R
T_LIFE = 20.0
DROP_GAP = 1.0
M0 = np.array([20000.0, 0.0, 2000.0], dtype=np.float64)
FY0 = np.array([17800.0, 0.0, 1800.0], dtype=np.float64)
TGT = np.array([0.0, 200.0, 5.0], dtype=np.float64)
DIST_M0 = float(np.linalg.norm(M0))
U_M = -M0 / DIST_M0
V_M = V_MISSILE * U_M
T_HIT = DIST_M0 / V_MISSILE


def device_info() -> str:
    if TORCH_OK and CUDA_OK:
        return f"torch-cuda:{torch.cuda.get_device_name(0)}"
    if TORCH_OK:
        return "torch-cpu"
    return "numpy-cpu"


# ====================== 几何：批量多云遮蔽 ======================
def kinematics_batch(theta, v, t_drop, tau):
    """theta,v,t_drop,tau: shape (N,) -> P_drop,P_det,t_det shape (N,3)/(N,)"""
    theta = np.asarray(theta, dtype=np.float64)
    v = np.asarray(v, dtype=np.float64)
    t_drop = np.asarray(t_drop, dtype=np.float64)
    tau = np.asarray(tau, dtype=np.float64)
    vx = v * np.cos(theta)
    vy = v * np.sin(theta)
    p_drop = np.stack(
        [FY0[0] + vx * t_drop, FY0[1] + vy * t_drop, np.full_like(t_drop, FY0[2])],
        axis=-1,
    )
    t_det = t_drop + tau
    p_det = np.stack(
        [
            p_drop[..., 0] + vx * tau,
            p_drop[..., 1] + vy * tau,
            p_drop[..., 2] - 0.5 * G * tau * tau,
        ],
        axis=-1,
    )
    return p_drop, p_det, t_det


def union_duration_numpy(P_det, t_det, n_time=8000, t_end=None):
    """
    P_det: (N, K, 3), t_det: (N, K)
    返回: dur (N,), 以及可选 per-bomb 粗时长 (N,K)
    判据：任一一枚云团在时刻 t 遮蔽 LOS 即有效；并集时长。
    """
    P_det = np.asarray(P_det, dtype=np.float64)
    t_det = np.asarray(t_det, dtype=np.float64)
    N, K, _ = P_det.shape
    if t_end is None:
        t_end = min(T_HIT, float(np.nanmax(t_det) + T_LIFE + 1.0))
    t0 = 0.0
    # 全局时间网格：覆盖所有可能遮蔽窗口
    t0 = max(0.0, float(np.nanmin(t_det)) - 0.5)
    t0 = max(0.0, t0)
    t1 = min(T_HIT, float(np.nanmax(t_det)) + T_LIFE)
    if t1 <= t0:
        return np.zeros(N), np.zeros((N, K))

    ts = np.linspace(t0, t1, n_time, dtype=np.float64)  # (T,)
    dt = ts[1] - ts[0]
    T = ts.shape[0]

    # 导弹轨迹 (T,3)
    M = M0[None, :] + V_M[None, :] * ts[:, None]
    AB = TGT[None, :] - M  # (T,3)
    L2 = np.sum(AB * AB, axis=1)  # (T,)
    L2 = np.maximum(L2, 1e-18)

    # 云团: (N,K,T,3)
    # C_x = P_det_x, C_z = P_det_z - vsink*(t-tdet) 仅当 t in [tdet, tdet+T_life]
    alive = (ts[None, None, :] >= t_det[:, :, None]) & (
        ts[None, None, :] <= t_det[:, :, None] + T_LIFE
    )  # (N,K,T)

    Cx = P_det[:, :, 0:1]  # (N,K,1)
    Cy = P_det[:, :, 1:2]
    Cz0 = P_det[:, :, 2:3]
    Cz = Cz0 - V_SINK * (ts[None, None, :] - t_det[:, :, None])

    # AC = C - M : broadcast (N,K,T,3)
    ACx = Cx - M[None, None, :, 0]
    ACy = Cy - M[None, None, :, 1]
    ACz = Cz - M[None, None, :, 2]

    # s = AC·AB / L2
    s = (ACx * AB[None, None, :, 0] + ACy * AB[None, None, :, 1] + ACz * AB[None, None, :, 2]) / L2[
        None, None, :
    ]
    sc = np.clip(s, 0.0, 1.0)
    px = M[None, None, :, 0] + sc * AB[None, None, :, 0]
    py = M[None, None, :, 1] + sc * AB[None, None, :, 1]
    pz = M[None, None, :, 2] + sc * AB[None, None, :, 2]
    d2 = (Cx - px) ** 2 + (Cy - py) ** 2 + (Cz - pz) ** 2
    bomb_ok = alive & (d2 <= R2) & (s >= 0.0) & (s <= 1.0)  # (N,K,T)

    any_ok = np.any(bomb_ok, axis=1)  # (N,T)
    dur = np.sum(any_ok, axis=1) * dt
    per = np.sum(bomb_ok, axis=2) * dt  # (N,K) 单弹时长（可重叠计）
    return dur.astype(np.float64), per.astype(np.float64)


def union_duration_torch(P_det, t_det, n_time=8000, device=None):
    """GPU/CPU torch 版批量并集时长。P_det (N,K,3), t_det (N,K)"""
    if device is None:
        device = "cuda" if CUDA_OK else "cpu"
    P_det = torch.as_tensor(P_det, dtype=torch.float32, device=device)
    t_det = torch.as_tensor(t_det, dtype=torch.float32, device=device)
    N, K, _ = P_det.shape

    t0 = float(torch.clamp(t_det.min() - 0.5, min=0.0).item())
    t1 = float(min(T_HIT, float(t_det.max().item()) + T_LIFE))
    if t1 <= t0:
        z = torch.zeros(N, device=device)
        return z, torch.zeros(N, K, device=device)

    ts = torch.linspace(t0, t1, n_time, device=device, dtype=torch.float32)
    dt = float(ts[1] - ts[0])
    M0_t = torch.tensor(M0, dtype=torch.float32, device=device)
    VM_t = torch.tensor(V_M, dtype=torch.float32, device=device)
    TGT_t = torch.tensor(TGT, dtype=torch.float32, device=device)

    M = M0_t[None, :] + VM_t[None, :] * ts[:, None]  # (T,3)
    AB = TGT_t[None, :] - M
    L2 = torch.clamp((AB * AB).sum(dim=1), min=1e-18)

    alive = (ts[None, None, :] >= t_det[:, :, None]) & (
        ts[None, None, :] <= t_det[:, :, None] + T_LIFE
    )
    Cx = P_det[:, :, 0:1]
    Cy = P_det[:, :, 1:2]
    Cz = P_det[:, :, 2:3] - V_SINK * (ts[None, None, :] - t_det[:, :, None])

    ACx = Cx - M[None, None, :, 0]
    ACy = Cy - M[None, None, :, 1]
    ACz = Cz - M[None, None, :, 2]
    s = (ACx * AB[None, None, :, 0] + ACy * AB[None, None, :, 1] + ACz * AB[None, None, :, 2]) / L2[
        None, None, :
    ]
    sc = torch.clamp(s, 0.0, 1.0)
    px = M[None, None, :, 0] + sc * AB[None, None, :, 0]
    py = M[None, None, :, 1] + sc * AB[None, None, :, 1]
    pz = M[None, None, :, 2] + sc * AB[None, None, :, 2]
    d2 = (Cx - px) ** 2 + (Cy - py) ** 2 + (Cz - pz) ** 2
    bomb_ok = alive & (d2 <= R2) & (s >= 0.0) & (s <= 1.0)
    any_ok = bomb_ok.any(dim=1)
    dur = any_ok.sum(dim=1).to(torch.float64) * dt
    per = bomb_ok.sum(dim=2).to(torch.float64) * dt
    return dur.detach().cpu().numpy(), per.detach().cpu().numpy()


def eval_strategies(X, n_time=8000, use_gpu=True):
    """
    X: (N, 8) = [theta, v, t1, tau1, t2, tau2, t3, tau3]
    返回 dur (N,), details list optional
    """
    X = np.asarray(X, dtype=np.float64)
    if X.ndim == 1:
        X = X[None, :]
    N = X.shape[0]
    theta = X[:, 0]
    v = X[:, 1]
    # 修复时间顺序：保证间隔
    t1 = X[:, 2]
    t2 = np.maximum(X[:, 4], t1 + DROP_GAP)
    t3 = np.maximum(X[:, 6], t2 + DROP_GAP)
    tau1, tau2, tau3 = X[:, 3], X[:, 5], X[:, 7]

    drops = []
    dets = []
    tds = []
    for ti, taui in ((t1, tau1), (t2, tau2), (t3, tau3)):
        pd, pdet, td = kinematics_batch(theta, v, ti, taui)
        drops.append(pd)
        dets.append(pdet)
        tds.append(td)
    P_drop = np.stack(drops, axis=1)  # (N,3,3)
    P_det = np.stack(dets, axis=1)
    t_det = np.stack(tds, axis=1)

    if use_gpu and TORCH_OK and CUDA_OK:
        # 分块避免显存爆
        chunk = 2048 if n_time <= 10000 else 1024
        durs = []
        pers = []
        for i0 in range(0, N, chunk):
            i1 = min(N, i0 + chunk)
            d, pe = union_duration_torch(P_det[i0:i1], t_det[i0:i1], n_time=n_time)
            durs.append(d)
            pers.append(pe)
        dur = np.concatenate(durs, axis=0)
        per = np.concatenate(pers, axis=0)
    else:
        # NumPy 也分块，控制内存 (N*K*T)
        # 约 N*3*8000*8B*~10 ≈ N*2MB 量级，N=4000 约 8GB 紧，故分块
        chunk = 512 if n_time >= 8000 else 1024
        durs = []
        pers = []
        for i0 in range(0, N, chunk):
            i1 = min(N, i0 + chunk)
            d, pe = union_duration_numpy(P_det[i0:i1], t_det[i0:i1], n_time=n_time)
            durs.append(d)
            pers.append(pe)
        dur = np.concatenate(durs, axis=0)
        per = np.concatenate(pers, axis=0)

    # 写回规范化后的时间
    Xn = X.copy()
    Xn[:, 2], Xn[:, 4], Xn[:, 6] = t1, t2, t3
    return dur, per, Xn, P_drop, P_det, t_det


def pack_strategy(x):
    x = np.asarray(x, dtype=np.float64).ravel()
    theta, v = float(x[0]), float(x[1])
    t1, tau1, t2, tau2, t3, tau3 = map(float, x[2:8])
    t2 = max(t2, t1 + DROP_GAP)
    t3 = max(t3, t2 + DROP_GAP)
    bombs = []
    for ti, taui in ((t1, tau1), (t2, tau2), (t3, tau3)):
        pd, pdet, td = kinematics_batch([theta], [v], [ti], [taui])
        bombs.append(
            {
                "t_drop": ti,
                "tau": taui,
                "t_det": float(td[0]),
                "P_drop": pd[0].tolist(),
                "P_det": pdet[0].tolist(),
            }
        )
    return {
        "theta": theta,
        "heading_deg": float(np.degrees(theta) % 360.0),
        "v": v,
        "bombs": bombs,
        "x": [theta, v, t1, tau1, t2, tau2, t3, tau3],
    }


def highres_union(x, n_time=200000, refine_edges=True):
    """高精度单策略并集时长 + 首段边界近似"""
    info = pack_strategy(x)
    P_det = np.array([b["P_det"] for b in info["bombs"]], dtype=np.float64)[None, :, :]
    t_det = np.array([b["t_det"] for b in info["bombs"]], dtype=np.float64)[None, :]
    if TORCH_OK and CUDA_OK:
        dur, per = union_duration_torch(P_det, t_det, n_time=n_time)
    else:
        dur, per = union_duration_numpy(P_det, t_det, n_time=n_time)
    return float(dur[0]), per[0], info


# ====================== 搜索 ======================
def sample_candidates(n_rand=8000, rng=None):
    if rng is None:
        rng = np.random.default_rng(2025)
    # 边界
    # theta 集中在 pi 附近 + 全周少量
    n1 = n_rand // 2
    n2 = n_rand - n1
    th1 = np.pi + (rng.random(n1) - 0.5) * 0.6
    th2 = rng.random(n2) * 2 * np.pi
    theta = np.concatenate([th1, th2])
    v = 70 + rng.random(n_rand) * 70
    # 投放：用间隔编码 d0,g1,g2 >=0, g>=1 通过 +1
    d0 = rng.random(n_rand) * 12.0  # first drop 0~12
    g1 = DROP_GAP + rng.random(n_rand) * 8.0
    g2 = DROP_GAP + rng.random(n_rand) * 8.0
    t1 = d0
    t2 = t1 + g1
    t3 = t2 + g2
    tau = 0.5 + rng.random((n_rand, 3)) * 10.0
    X = np.column_stack([theta, v, t1, tau[:, 0], t2, tau[:, 1], t3, tau[:, 2]])
    return X


def grid_candidates():
    """围绕 Q2 优解的结构化网格（可控规模）"""
    # Q2 附近 + 多弹时间轴
    ths = np.pi + np.linspace(-0.25, 0.25, 9)
    vs = np.array([70, 80, 90, 100, 110, 120, 130, 140], dtype=float)
    # 三弹投放时刻模式
    patterns = [
        (0.0, 1.0, 2.0),
        (0.0, 1.2, 2.5),
        (0.0, 1.5, 3.0),
        (0.0, 2.0, 4.0),
        (0.2, 1.5, 3.0),
        (0.5, 2.0, 4.0),
        (1.0, 2.5, 4.5),
        (0.0, 1.0, 3.5),
        (0.0, 1.0, 5.0),
        (0.0, 2.5, 5.0),
        (0.5, 1.5, 2.5),
        (0.0, 1.5, 4.5),
    ]
    taus_sets = [
        (2.0, 2.5, 3.0),
        (2.5, 2.5, 2.5),
        (2.5, 3.0, 3.5),
        (3.0, 3.5, 4.0),
        (2.0, 3.0, 4.0),
        (1.5, 2.5, 3.5),
        (3.5, 3.5, 3.5),
        (2.5, 4.0, 5.0),
        (1.8, 2.8, 3.8),
        (2.2, 3.2, 4.2),
    ]
    rows = []
    for th in ths:
        for v in vs:
            for (t1, t2, t3) in patterns:
                for (a, b, c) in taus_sets:
                    rows.append([th, v, t1, a, t2, b, t3, c])
    # Q2 优解扩展：固定航向速度，扫三弹
    q2_th, q2_v = 3.0880757565, 71.8890217683
    for t1 in np.linspace(0.0, 3.0, 8):
        for g1 in np.linspace(1.0, 5.0, 7):
            for g2 in np.linspace(1.0, 5.0, 7):
                for tau in np.linspace(1.5, 6.0, 8):
                    # 共用 tau 或略变
                    rows.append([q2_th, q2_v, t1, tau, t1 + g1, tau + 0.3, t1 + g1 + g2, tau + 0.6])
                    rows.append([q2_th, q2_v, t1, tau, t1 + g1, tau, t1 + g1 + g2, tau])
    # 问题1 风格
    rows.append([np.pi, 120, 1.5, 3.6, 2.5, 3.6, 3.5, 3.6])
    rows.append([np.pi, 100, 0.0, 2.5, 1.0, 2.5, 2.0, 2.5])
    rows.append([q2_th, q2_v, 0.0, 2.5, 1.0, 2.5, 2.0, 2.5])
    return np.asarray(rows, dtype=np.float64)


def local_refine(x0, n_time=12000, steps=60, spans=None, use_gpu=True):
    if spans is None:
        spans = np.array([0.06, 6.0, 0.8, 0.8, 0.8, 0.8, 0.8, 0.8], dtype=float)
    lb = np.array([0.0, 70.0, 0.0, 0.3, 0.0, 0.3, 0.0, 0.3])
    ub = np.array([2 * np.pi, 140.0, 20.0, 14.0, 30.0, 14.0, 40.0, 14.0])
    x = np.array(x0, dtype=float)
    # normalize gaps
    x[4] = max(x[4], x[2] + DROP_GAP)
    x[6] = max(x[6], x[4] + DROP_GAP)
    best = float(eval_strategies(x[None, :], n_time=n_time, use_gpu=use_gpu)[0][0])
    span = spans.copy()
    rng = np.random.default_rng(7)
    for _ in range(steps):
        improved = False
        # 坐标轮换
        for j in range(8):
            for sgn in (-1.0, 1.0):
                trial = x.copy()
                trial[j] += sgn * span[j]
                trial = np.minimum(np.maximum(trial, lb), ub)
                trial[0] %= 2 * np.pi
                trial[4] = max(trial[4], trial[2] + DROP_GAP)
                trial[6] = max(trial[6], trial[4] + DROP_GAP)
                val = float(eval_strategies(trial[None, :], n_time=n_time, use_gpu=use_gpu)[0][0])
                if val > best + 1e-8:
                    best, x = val, trial
                    improved = True
        # 批量随机扰动
        noise = (rng.random((24, 8)) - 0.5) * 2.0 * span[None, :]
        trials = x[None, :] + noise
        trials = np.minimum(np.maximum(trials, lb), ub)
        trials[:, 0] %= 2 * np.pi
        trials[:, 4] = np.maximum(trials[:, 4], trials[:, 2] + DROP_GAP)
        trials[:, 6] = np.maximum(trials[:, 6], trials[:, 4] + DROP_GAP)
        vals = eval_strategies(trials, n_time=n_time, use_gpu=use_gpu)[0]
        imax = int(np.argmax(vals))
        if vals[imax] > best + 1e-8:
            best = float(vals[imax])
            x = trials[imax]
            improved = True
        if not improved:
            span *= 0.6
            if np.all(span < 1e-3):
                break
        else:
            span = np.minimum(span * 1.05, spans * 1.2)
    return x, best


def write_result1_xlsx(info, per_bomb, total_dur, path: Path):
    try:
        import openpyxl
    except ImportError:
        # 纯 csv 兜底
        path = path.with_suffix(".csv")
        with open(path, "w", encoding="utf-8-sig") as f:
            f.write(
                "无人机运动方向,无人机运动速度 (m/s),烟幕干扰弹编号,"
                "投放x,投放y,投放z,起爆x,起爆y,起爆z,有效干扰时长 (s)\n"
            )
            for i, b in enumerate(info["bombs"]):
                f.write(
                    f"{info['heading_deg']:.6f},{info['v']:.6f},{i+1},"
                    f"{b['P_drop'][0]:.6f},{b['P_drop'][1]:.6f},{b['P_drop'][2]:.6f},"
                    f"{b['P_det'][0]:.6f},{b['P_det'][1]:.6f},{b['P_det'][2]:.6f},"
                    f"{float(per_bomb[i]):.6f}\n"
                )
        return path

    src = ATTACH / "result1.xlsx"
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    # 表头已在模板；填 3 行
    # 方向/速度：三行相同（模板注释：以x正向逆时针 0~360）
    for i, b in enumerate(info["bombs"]):
        r = 2 + i
        ws.cell(r, 1, round(info["heading_deg"], 6))
        ws.cell(r, 2, round(info["v"], 6))
        ws.cell(r, 3, i + 1)
        for j in range(3):
            ws.cell(r, 4 + j, round(b["P_drop"][j], 6))
            ws.cell(r, 7 + j, round(b["P_det"][j], 6))
        # 单弹有效时长；总并集写在备注或第一行额外说明
        ws.cell(r, 10, round(float(per_bomb[i]), 6))
    # 在第5行写总并集说明
    ws.cell(5, 1, f"三枚并集有效遮蔽总时长 = {total_dur:.6f} s")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    # 同步到附件目录
    try:
        shutil.copy2(path, ATTACH / "result1.xlsx")
    except Exception:
        pass
    return path


def main():
    t_all = time.perf_counter()
    print("==== 问题3：FY1 三弹干扰 M1 ====")
    print(f"backend = {device_info()}")
    print(f"t_hit = {T_HIT:.6f} s")
    use_gpu = CUDA_OK
    if use_gpu:
        print(f"GPU: {torch.cuda.get_device_name(0)}")
        # 预热
        _ = eval_strategies(np.array([[np.pi, 120, 0, 2.5, 1, 2.5, 2, 2.5]]), n_time=2000, use_gpu=True)
        torch.cuda.synchronize()

    # 基准：三枚复制 Q2 单弹策略（间隔1s）
    x_q2 = np.array([3.0880757565, 71.8890217683, 0.0, 2.5032397513, 1.0, 2.5032397513, 2.0, 2.5032397513])
    d0, p0, _, _, _, _ = eval_strategies(x_q2[None, :], n_time=20000, use_gpu=use_gpu)
    print(f"[基准 Q2×3] union≈{d0[0]:.4f} s  per={p0[0]}")

    # ---- 粗搜 ----
    print("\n---- 粗搜 ----")
    Xg = grid_candidates()
    Xr = sample_candidates(10000)
    X = np.vstack([x_q2[None, :], Xg, Xr])
    print(f"候选数 {len(X)}  grid={len(Xg)} rand={len(Xr)}")
    n_coarse = 6000
    t0 = time.perf_counter()
    durs, pers, Xn, _, _, _ = eval_strategies(X, n_time=n_coarse, use_gpu=use_gpu)
    if use_gpu:
        torch.cuda.synchronize()
    print(f"粗搜完成 {time.perf_counter()-t0:.2f}s  best={durs.max():.4f}")
    order = np.argsort(-durs)
    print("Top-12:")
    for k in range(12):
        i = order[k]
        x = Xn[i]
        print(
            f"  #{k+1:2d} {durs[i]:.4f} th={np.degrees(x[0])%360:.2f} v={x[1]:.1f} "
            f"t=[{x[2]:.2f},{x[4]:.2f},{x[6]:.2f}] tau=[{x[3]:.2f},{x[5]:.2f},{x[7]:.2f}]"
        )

    # ---- 精修 ----
    n_top = 25
    print(f"\n---- 局部精修 Top-{n_top} ----")
    refined = []
    t1 = time.perf_counter()
    for k in range(n_top):
        x0 = Xn[order[k]]
        xf, df = local_refine(x0, n_time=10000, steps=55, use_gpu=use_gpu)
        refined.append((df, xf))
        print(f"  {k+1}/{n_top}: {durs[order[k]]:.4f} -> {df:.4f}")
    refined.sort(key=lambda z: -z[0])
    best_mid, x_best = refined[0]
    print(f"精修最优 {best_mid:.6f}  耗时 {time.perf_counter()-t1:.1f}s")

    # 二次精修
    x_best, best2 = local_refine(
        x_best,
        n_time=20000,
        steps=40,
        spans=np.array([0.03, 3.0, 0.4, 0.4, 0.4, 0.4, 0.4, 0.4]),
        use_gpu=use_gpu,
    )
    print(f"二次精修 {best2:.6f}")

    # 高精度
    total, per, info = highres_union(x_best, n_time=250000)
    print("\n==== 问题3 最优策略 ====")
    print(f"并集有效遮蔽时长 = {total:.6f} s")
    print(f"单弹时长(可重叠) = {per}")
    print(f"航向 = {info['heading_deg']:.6f} deg, v={info['v']:.6f} m/s")
    for i, b in enumerate(info["bombs"]):
        print(
            f"  弹{i+1}: t_drop={b['t_drop']:.4f} tau={b['tau']:.4f} t_det={b['t_det']:.4f} "
            f"P_drop={np.round(b['P_drop'],3)} P_det={np.round(b['P_det'],3)}"
        )

    elapsed = time.perf_counter() - t_all

    # 写文件
    txt = OUT / "q3_result.txt"
    with open(txt, "w", encoding="utf-8") as f:
        f.write("==== 2025国赛A题 问题3 结果 ====\n")
        f.write("FY1 投放 3 枚烟幕干扰弹，干扰 M1；遮蔽取并集（可不连续）\n")
        f.write(f"backend = {device_info()}\n")
        f.write(f"并集有效遮蔽时长 = {total:.6f} s\n")
        f.write(f"单弹有效时长(各自，可重叠) = [{per[0]:.6f}, {per[1]:.6f}, {per[2]:.6f}] s\n")
        f.write("\n--- 无人机 ---\n")
        f.write(f"航向角 theta = {info['theta']:.10f} rad = {info['heading_deg']:.6f} deg\n")
        f.write(f"飞行速度 v   = {info['v']:.10f} m/s\n")
        for i, b in enumerate(info["bombs"]):
            f.write(f"\n--- 烟幕干扰弹 {i+1} ---\n")
            f.write(f"投放时刻 t_drop = {b['t_drop']:.10f} s\n")
            f.write(f"引信延时 tau    = {b['tau']:.10f} s\n")
            f.write(f"起爆时刻 t_det  = {b['t_det']:.10f} s\n")
            f.write(
                f"投放点 P_drop = [{b['P_drop'][0]:.10f}, {b['P_drop'][1]:.10f}, {b['P_drop'][2]:.10f}]\n"
            )
            f.write(
                f"起爆点 P_det  = [{b['P_det'][0]:.10f}, {b['P_det'][1]:.10f}, {b['P_det'][2]:.10f}]\n"
            )
            f.write(f"单弹有效时长 = {per[i]:.6f} s\n")
        f.write(f"\n总运行时间 = {elapsed:.2f} s\n")
        f.write(f"决策向量 x = {info['x']}\n")

    with open(OUT / "q3_result.json", "w", encoding="utf-8") as f:
        json.dump(
            {
                "total_union": total,
                "per_bomb": per.tolist(),
                "info": info,
                "elapsed_s": elapsed,
                "backend": device_info(),
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    xlsx_path = write_result1_xlsx(info, per, total, OUT / "result1.xlsx")
    print(f"\n结果: {display_path(txt)}")
    print(f"Excel: {display_path(xlsx_path)}")
    print(f"总耗时 {elapsed:.1f}s")


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
问题4：圆柱严格全遮蔽下，FY1/FY2/FY3 各投 1 枚烟幕干扰弹，对 M1 并集遮蔽最大化

策略：
  1) 每机先独立优化单弹（覆盖早期/中期/晚期拦截窗口）
  2) 用三机独立最优作为 warm-start 联合坐标下降精修
  3) 随机扰动 + 时间错开搜索，防止落回局部最优
"""
from __future__ import annotations

import json
import shutil
import sys
import time
from pathlib import Path

import numpy as np

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
from common.smoke_geom import (  # noqa: E402
    PTS_FAST,
    PTS_FINE,
    PTS_OPT,
    T_HIT,
    UAVS,
    device_info,
    heading_deg,
    kinematics,
    kinematics_batch,
    shield_duration,
    union_duration_batch,
    union_duration_single_path,
)

OUT = ROOT / "结果"

def display_path(path):
    return Path(path).resolve().relative_to(ROOT.parent.resolve()).as_posix()

OUT.mkdir(exist_ok=True)
ATTACH = ROOT.parent / "附件"

MODE = "cylinder"
PTS_COARSE = PTS_FAST
PTS_MID = PTS_OPT
PTS_HI = PTS_FINE

UAV_NAMES = ["FY1", "FY2", "FY3"]
FY_POS = [UAVS[n] for n in UAV_NAMES]


def aim_theta(fy0, target_xy=(0.0, 0.0)):
    dxy = np.array(target_xy, dtype=float) - np.asarray(fy0, dtype=float)[:2]
    return float(np.arctan2(dxy[1], dxy[0]))


def pack_uav(name, theta, v, t_drop, tau):
    fy0 = UAVS[name]
    v_fy, p_drop, t_det, p_det = kinematics(theta, v, t_drop, tau, fy0=fy0)
    return {
        "name": name,
        "theta": float(theta),
        "heading_deg": heading_deg(theta),
        "v": float(v),
        "t_drop": float(t_drop),
        "tau": float(tau),
        "t_det": float(t_det),
        "P_drop": p_drop.tolist(),
        "P_det": p_det.tolist(),
        "v_fy": v_fy.tolist(),
        "fy0": fy0.tolist(),
    }


def clip_x(x):
    x = np.asarray(x, dtype=float).copy()
    for i in range(3):
        x[4 * i + 0] = np.mod(x[4 * i + 0], 2 * np.pi)
        x[4 * i + 1] = np.clip(x[4 * i + 1], 70.0, 140.0)
        x[4 * i + 2] = max(0.0, min(50.0, float(x[4 * i + 2])))
        x[4 * i + 3] = max(0.2, min(16.0, float(x[4 * i + 3])))
    return x


def eval_batch(X, n_time=5000, use_gpu=True, pts=None):
    X = np.asarray(X, dtype=np.float64)
    if X.ndim == 1:
        X = X[None, :]
    if pts is None:
        pts = PTS_COARSE
    dets = []
    tds = []
    for i, fy0 in enumerate(FY_POS):
        th = X[:, 4 * i + 0]
        v = np.clip(X[:, 4 * i + 1], 70.0, 140.0)
        td = np.clip(X[:, 4 * i + 2], 0.0, 50.0)
        tau = np.clip(X[:, 4 * i + 3], 0.2, 16.0)
        _, pdet, tdet = kinematics_batch(th, v, td, tau, fy0=fy0)
        dets.append(pdet)
        tds.append(tdet)
    P_det = np.stack(dets, axis=1)
    t_det = np.stack(tds, axis=1)
    dur, per = union_duration_batch(
        P_det, t_det, mode=MODE, pts=pts, n_time=n_time, use_gpu=use_gpu, missile="M1"
    )
    return dur, per, P_det, t_det


def highres(x, pts=None, n=220000):
    if pts is None:
        pts = PTS_HI
    x = clip_x(x)
    bombs = []
    infos = []
    for i, name in enumerate(UAV_NAMES):
        info = pack_uav(name, x[4 * i], x[4 * i + 1], x[4 * i + 2], x[4 * i + 3])
        infos.append(info)
        bombs.append((np.array(info["P_det"]), info["t_det"]))
    total, per, segs = union_duration_single_path(
        bombs, mode=MODE, pts=pts, n=n, refine=True, missile="M1"
    )
    return total, per, segs, infos


def optimize_single_uav(name, n_rand=2800, use_gpu=True, warm=None, seed=0):
    """独立优化单机对 M1 的单弹遮蔽。"""
    fy0 = UAVS[name]
    th0 = aim_theta(fy0)
    thm = aim_theta(fy0, (20000.0, 0.0))
    tht = aim_theta(fy0, (0.0, 200.0))
    rng = np.random.default_rng(seed + abs(hash(name)) % 10000)

    rows = []
    if warm is not None:
        rows.append(list(warm))
    # FY1 Q2 cylinder
    if name == "FY1":
        rows.append([3.0825785063, 70.0, 0.0, 2.4841194504])
        rows.append([3.129, 140.0, 0.0, 3.511])
    # time-window priors by UAV geometry
    if name == "FY1":
        td_grid = [0.0, 0.5, 1.0, 2.0, 3.0, 5.0]
        tau_grid = [1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 5.0]
    elif name == "FY2":
        td_grid = [2.0, 4.0, 6.0, 8.0, 10.0, 12.0, 15.0, 18.0]
        tau_grid = [2.0, 3.0, 4.0, 5.0, 6.0, 8.0]
    else:  # FY3 farther, later window
        td_grid = [8.0, 12.0, 16.0, 20.0, 24.0, 28.0, 32.0, 36.0]
        tau_grid = [3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 10.0]

    th_list = (
        [th0, thm, tht, np.pi, 0.0]
        + list(th0 + np.linspace(-1.2, 1.2, 9))
        + list(thm + np.linspace(-0.8, 0.8, 7))
        + list(tht + np.linspace(-0.8, 0.8, 7))
    )
    for th in th_list:
        for v in [70, 90, 110, 130, 140]:
            for td in td_grid:
                for tau in tau_grid:
                    rows.append([th, v, td, tau])

    Xr = np.column_stack(
        [
            (th0 + (rng.random(n_rand) - 0.5) * 2.4) % (2 * np.pi),
            70 + rng.random(n_rand) * 70,
            rng.choice(td_grid, size=n_rand)
            + (rng.random(n_rand) - 0.5) * 3.0,
            0.5 + rng.random(n_rand) * 11,
        ]
    )
    X = np.vstack([np.asarray(rows, float), Xr])
    X[:, 0] %= 2 * np.pi
    X[:, 1] = np.clip(X[:, 1], 70, 140)
    X[:, 2] = np.clip(X[:, 2], 0, 50)
    X[:, 3] = np.clip(X[:, 3], 0.3, 16)
    X = np.unique(np.round(X, 5), axis=0)

    best = -1.0
    bestx = None
    chunk = 900
    for i0 in range(0, len(X), chunk):
        i1 = min(len(X), i0 + chunk)
        th, v, td, tau = X[i0:i1, 0], X[i0:i1, 1], X[i0:i1, 2], X[i0:i1, 3]
        _, P, Td = kinematics_batch(th, v, td, tau, fy0=fy0)
        d, _ = union_duration_batch(
            P[:, None, :],
            Td[:, None],
            mode=MODE,
            pts=PTS_COARSE,
            n_time=3500,
            use_gpu=use_gpu,
            missile="M1",
        )
        j = int(np.argmax(d))
        if d[j] > best:
            best = float(d[j])
            bestx = X[i0 + j].copy()

    # local refine
    lb = np.array([0.0, 70.0, 0.0, 0.3])
    ub = np.array([2 * np.pi, 140.0, 50.0, 16.0])
    x = bestx.copy()
    val = best
    span = np.array([0.1, 10.0, 1.5, 1.5])
    for _ in range(36):
        improved = False
        for d in range(4):
            for sgn in (-1.0, 1.0):
                trial = x.copy()
                trial[d] += sgn * span[d]
                trial[0] %= 2 * np.pi
                trial = np.clip(trial, lb, ub)
                _, pd, td_ = kinematics_batch(
                    [trial[0]], [trial[1]], [trial[2]], [trial[3]], fy0=fy0
                )
                dv, _ = union_duration_batch(
                    pd[:, None, :],
                    td_[:, None],
                    mode=MODE,
                    pts=PTS_MID,
                    n_time=4000,
                    use_gpu=use_gpu,
                    missile="M1",
                )
                if dv[0] > val + 1e-9:
                    val = float(dv[0])
                    x = trial
                    improved = True
        noise = (rng.random((36, 4)) - 0.5) * 2 * span
        trials = np.clip(x[None, :] + noise, lb, ub)
        trials[:, 0] %= 2 * np.pi
        _, pd, td_ = kinematics_batch(
            trials[:, 0], trials[:, 1], trials[:, 2], trials[:, 3], fy0=fy0
        )
        dvs, _ = union_duration_batch(
            pd[:, None, :],
            td_[:, None],
            mode=MODE,
            pts=PTS_MID,
            n_time=3500,
            use_gpu=use_gpu,
            missile="M1",
        )
        j = int(np.argmax(dvs))
        if dvs[j] > val + 1e-9:
            val = float(dvs[j])
            x = trials[j]
            improved = True
        span *= 0.9 if improved else 1.05
        span = np.minimum(span, np.array([0.15, 15, 2.5, 2.5]))

    v_fy, p_drop, t_det, p_det = kinematics(*x, fy0=fy0)
    dur_hi, segs = shield_duration(
        p_det, t_det, mode=MODE, pts=PTS_MID, n=80000, refine=True, missile="M1"
    )
    print(
        f"  {name} solo: {dur_hi:.4f}s  heading={heading_deg(x[0]):.2f} "
        f"v={x[1]:.2f} td={x[2]:.3f} tau={x[3]:.3f} t_det={t_det:.3f} segs={segs}"
    )
    return x, float(dur_hi)


def local_refine(x0, n_time=5000, steps=45, spans=None, use_gpu=True, pts=None):
    if spans is None:
        spans = np.array([0.08, 8.0, 1.2, 1.2] * 3)
    if pts is None:
        pts = PTS_MID
    lb = np.array([0.0, 70.0, 0.0, 0.3] * 3)
    ub = np.array([2 * np.pi, 140.0, 50.0, 16.0] * 3)
    x = clip_x(x0)
    best = float(eval_batch(x[None, :], n_time=n_time, use_gpu=use_gpu, pts=pts)[0][0])
    span = spans.copy()
    rng = np.random.default_rng(2025)
    for it in range(steps):
        improved = False
        for d in range(12):
            for sgn in (-1.0, 1.0):
                trial = x.copy()
                trial[d] += sgn * span[d]
                trial = clip_x(np.clip(trial, lb, ub))
                val = float(
                    eval_batch(trial[None, :], n_time=n_time, use_gpu=use_gpu, pts=pts)[0][0]
                )
                if val > best + 1e-9:
                    best = val
                    x = trial
                    improved = True
        noise = (rng.random((64, 12)) - 0.5) * 2.0 * span[None, :]
        trials = np.array([clip_x(np.clip(x + n, lb, ub)) for n in noise])
        vals = eval_batch(trials, n_time=n_time, use_gpu=use_gpu, pts=pts)[0]
        j = int(np.argmax(vals))
        if vals[j] > best + 1e-9:
            best = float(vals[j])
            x = trials[j].copy()
            improved = True
        if improved:
            span *= 0.90
        else:
            span = np.minimum(span * 1.06, spans * 1.3)
        if (it + 1) % 10 == 0:
            print(f"    joint refine it={it+1} best={best:.4f}")
    return x, best


def build_joint_candidates(solo_xs, rng):
    """组合独立最优 + 时间错开扰动。"""
    rows = []
    # pure concat of solo optima
    x0 = np.concatenate(solo_xs)
    rows.append(x0)

    # swap/permutations of timing styles
    for scale in [0.8, 1.0, 1.2]:
        x = x0.copy()
        for i in range(3):
            x[4 * i + 2] = max(0.0, solo_xs[i][2] * scale)
            x[4 * i + 3] = np.clip(solo_xs[i][3] * (0.9 + 0.1 * scale), 0.3, 16)
        rows.append(x)

    # random perturbations around solo
    for _ in range(2500):
        x = x0.copy()
        for i in range(3):
            x[4 * i + 0] = (x[4 * i + 0] + (rng.random() - 0.5) * 0.5) % (2 * np.pi)
            x[4 * i + 1] = np.clip(x[4 * i + 1] + (rng.random() - 0.5) * 30, 70, 140)
            x[4 * i + 2] = max(0.0, x[4 * i + 2] + (rng.random() - 0.5) * 4)
            x[4 * i + 3] = np.clip(x[4 * i + 3] + (rng.random() - 0.5) * 2.5, 0.3, 16)
        rows.append(x)

    # extra: force staggered times even if solo times collide
    # FY1 early, FY2 mid, FY3 late
    th1, v1 = solo_xs[0][0], solo_xs[0][1]
    th2, v2 = solo_xs[1][0], solo_xs[1][1]
    th3, v3 = solo_xs[2][0], solo_xs[2][1]
    for t1 in [0.0, 0.5, 1.0]:
        for tau1 in [2.0, 2.5, 3.5]:
            for t2 in [4.0, 6.0, 8.0, 10.0, 12.0]:
                for tau2 in [3.0, 4.5, 6.0]:
                    for t3 in [18.0, 22.0, 26.0, 30.0, 34.0]:
                        for tau3 in [4.0, 6.0, 8.0]:
                            rows.append(
                                [th1, v1, t1, tau1, th2, v2, t2, tau2, th3, v3, t3, tau3]
                            )

    # broader random with staggered priors
    n_rand = 4000
    Xr = np.zeros((n_rand, 12))
    priors_td = [
        (0.0, 5.0),
        (3.0, 18.0),
        (15.0, 40.0),
    ]
    for i, fy0 in enumerate(FY_POS):
        th0 = aim_theta(fy0)
        lo, hi = priors_td[i]
        Xr[:, 4 * i + 0] = (th0 + (rng.random(n_rand) - 0.5) * 1.8) % (2 * np.pi)
        Xr[:, 4 * i + 1] = 70 + rng.random(n_rand) * 70
        Xr[:, 4 * i + 2] = lo + rng.random(n_rand) * (hi - lo)
        Xr[:, 4 * i + 3] = 0.5 + rng.random(n_rand) * 11

    X = np.vstack([np.asarray(rows, float), Xr])
    X = np.array([clip_x(r) for r in X])
    X = np.unique(np.round(X, 5), axis=0)
    return X


def write_result2_xlsx(infos, per, total, path: Path):
    template = ATTACH / "result2.xlsx"
    if load_workbook is not None and template.exists():
        wb = load_workbook(template)
        ws = wb.active
    else:
        if Workbook is None:
            raise RuntimeError("openpyxl required")
        wb = Workbook()
        ws = wb.active
        headers = [
            "无人机编号",
            "无人机运动方向",
            "无人机运动速度 (m/s)",
            "烟幕干扰弹投放点的x坐标 (m)",
            "烟幕干扰弹投放点的y坐标 (m)",
            "烟幕干扰弹投放点的z坐标 (m)",
            "烟幕干扰弹起爆点的x坐标 (m)",
            "烟幕干扰弹起爆点的y坐标 (m)",
            "烟幕干扰弹起爆点的z坐标 (m)",
            "有效干扰时长 (s)",
        ]
        for j, h in enumerate(headers, 1):
            ws.cell(1, j, h)

    for i, info in enumerate(infos):
        row = i + 2
        ws.cell(row, 1, info["name"])
        ws.cell(row, 2, float(info["heading_deg"]))
        ws.cell(row, 3, float(info["v"]))
        for j in range(3):
            ws.cell(row, 4 + j, float(info["P_drop"][j]))
            ws.cell(row, 7 + j, float(info["P_det"][j]))
        ws.cell(row, 10, float(per[i]))
    ws.cell(5, 1, f"三机并集有效遮蔽总时长 = {total:.6f} s （圆柱严格全遮蔽）")
    ws.cell(6, 2, "注：以x轴为正向，逆时针方向为正，取值0~360（度）。")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    try:
        shutil.copy2(path, ATTACH / "result2.xlsx")
    except Exception as e:
        print("copy to 附件 failed:", e)


def main():
    t0 = time.perf_counter()
    use_gpu = True
    print("==== 问题4 圆柱严格全遮蔽 FY1/2/3 各1弹 vs M1 ====")
    print("device:", device_info())
    print(f"t_hit={T_HIT:.6f}")
    print(f"pts coarse/mid/hi = {PTS_COARSE.shape[0]}/{PTS_MID.shape[0]}/{PTS_HI.shape[0]}")

    print("---- 阶段1: 单机独立最优 ----")
    solo_xs = []
    solo_ds = []
    for name in UAV_NAMES:
        x, d = optimize_single_uav(name, n_rand=2500, use_gpu=use_gpu)
        solo_xs.append(x)
        solo_ds.append(d)
    x_concat = clip_x(np.concatenate(solo_xs))
    d_concat = float(
        eval_batch(x_concat[None, :], n_time=8000, use_gpu=use_gpu, pts=PTS_MID)[0][0]
    )
    print(f"独立最优拼接并集(mid) = {d_concat:.4f}  (solo sum={sum(solo_ds):.4f})")

    print("---- 阶段2: 联合粗搜 ----")
    rng = np.random.default_rng(2025)
    X = build_joint_candidates(solo_xs, rng)
    n_all = len(X)
    print(f"联合候选 {n_all}")
    t_c = time.perf_counter()
    chunk = 500
    durs = np.zeros(n_all)
    for i0 in range(0, n_all, chunk):
        i1 = min(n_all, i0 + chunk)
        durs[i0:i1], _, _, _ = eval_batch(
            X[i0:i1], n_time=3500, use_gpu=use_gpu, pts=PTS_COARSE
        )
        if i1 == n_all or (i1 // chunk) % 5 == 0:
            print(
                f"  coarse {i1}/{n_all} best={durs[:i1].max():.4f} "
                f"({time.perf_counter()-t_c:.1f}s)"
            )

    # inject concat
    d_cat_c = float(
        eval_batch(x_concat[None, :], n_time=3500, use_gpu=use_gpu, pts=PTS_COARSE)[0][0]
    )
    X = np.vstack([X, x_concat[None, :]])
    durs = np.concatenate([durs, [d_cat_c]])

    order = np.argsort(-durs)
    Xs, ds = X[order], durs[order]
    print("Top-8 联合粗搜:")
    for k in range(min(8, len(ds))):
        xk = Xs[k]
        print(
            f"  #{k+1:2d} {ds[k]:.4f} "
            f"FY1 td={xk[2]:.2f}/tau={xk[3]:.2f} | "
            f"FY2 td={xk[6]:.2f}/tau={xk[7]:.2f} | "
            f"FY3 td={xk[10]:.2f}/tau={xk[11]:.2f}"
        )

    print("---- 阶段3: 联合精修 ----")
    n_top = 8
    refined = []
    t_r = time.perf_counter()
    for k in range(min(n_top, len(Xs))):
        xf, df = local_refine(
            Xs[k],
            n_time=4500,
            steps=30,
            use_gpu=use_gpu,
            pts=PTS_MID,
            spans=np.array([0.06, 6.0, 1.0, 1.0] * 3),
        )
        d_hi = float(eval_batch(xf[None, :], n_time=9000, use_gpu=use_gpu, pts=PTS_MID)[0][0])
        refined.append((d_hi, xf, df))
        print(f"  {k+1}/{n_top}: {ds[k]:.4f} -> {df:.4f} / hi={d_hi:.4f}")
    refined.sort(key=lambda t: -t[0])
    x_best = refined[0][1]
    print(f"精修最优 {refined[0][0]:.6f}  耗时 {time.perf_counter()-t_r:.1f}s")

    # final polish
    x_best, _ = local_refine(
        x_best,
        n_time=8000,
        steps=25,
        use_gpu=use_gpu,
        pts=PTS_MID,
        spans=np.array([0.03, 3.0, 0.5, 0.5] * 3),
    )

    total, per, segs, infos = highres(x_best, pts=PTS_HI, n=220000)
    print("\n==== 问题4 圆柱最优 ====")
    print(f"并集时长 = {total:.6f} s")
    print(f"单弹 = {per}")
    print(f"分段 = {segs}")
    for i, info in enumerate(infos):
        print(
            f"  {info['name']}: heading={info['heading_deg']:.4f} v={info['v']:.4f} "
            f"t_drop={info['t_drop']:.4f} tau={info['tau']:.4f} per={per[i]:.4f}"
        )

    elapsed = time.perf_counter() - t0

    def conv(o):
        if isinstance(o, (np.floating,)):
            return float(o)
        if isinstance(o, (np.integer,)):
            return int(o)
        if isinstance(o, np.ndarray):
            return o.tolist()
        if isinstance(o, list):
            return [conv(x) for x in o]
        if isinstance(o, tuple):
            return [conv(x) for x in o]
        if isinstance(o, dict):
            return {k: conv(v) for k, v in o.items()}
        return o

    result = {
        "total_union": total,
        "per_bomb": per,
        "segments": segs,
        "uavs": infos,
        "x": x_best.tolist(),
        "solo_durations": solo_ds,
        "mode": "cylinder_strict_full_cover",
        "device": device_info(),
        "elapsed_s": elapsed,
    }
    with open(OUT / "q4_cylinder_result.json", "w", encoding="utf-8") as f:
        json.dump(conv(result), f, ensure_ascii=False, indent=2)

    with open(OUT / "q4_cylinder_result.txt", "w", encoding="utf-8") as f:
        f.write("==== 2025国赛A题 问题4 圆柱严格全遮蔽 结果 ====\n")
        f.write("判据: 云团到导弹-圆柱表面采样点视线段距离均 <= 10 m\n")
        f.write(f"并集有效遮蔽时长 = {total:.6f} s\n")
        f.write(f"分段 = {segs}\n")
        f.write(f"单弹时长 = {per}\n")
        for info, d in zip(infos, per):
            f.write(
                f"{info['name']}: heading={info['heading_deg']:.10f} deg, v={info['v']:.10f}, "
                f"t_drop={info['t_drop']:.10f}, tau={info['tau']:.10f}, t_det={info['t_det']:.10f}\n"
            )
            f.write(f"  P_drop={info['P_drop']}\n")
            f.write(f"  P_det={info['P_det']}\n")
            f.write(f"  单弹时长={d:.6f}\n")
        f.write(f"device={device_info()}\n")
        f.write(f"elapsed={elapsed:.2f}s\n")

    xlsx_path = OUT / "result2.xlsx"
    write_result2_xlsx(infos, per, total, xlsx_path)
    print(f"已写 {display_path(xlsx_path)} 与 附件/result2.xlsx")
    print(f"总耗时 {elapsed:.1f}s")


if __name__ == "__main__":
    main()
